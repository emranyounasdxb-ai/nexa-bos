from __future__ import annotations

import asyncio
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from uuid import UUID

import pytest
from helpers import authenticate, owner_client, spawned_client, unique_tag
from httpx import AsyncClient
from nexa_bos_api.applications.models import Application
from nexa_bos_api.finance.api import router as finance_router
from nexa_bos_api.finance.models import (
    FinanceComponent,
    FinancePayout,
    FinancePayoutPeriod,
    FinancePeriodTransition,
)
from nexa_bos_api.identity.models import AuditEvent
from nexa_bos_api.main import app
from openpyxl import load_workbook
from sqlalchemy import func, select
from test_applications import _catalog, _create_app, _customer
from test_reports import _reporting_user


def _unique_month() -> date:
    value = int(unique_tag(), 16)
    return date(2200 + (value // 12) % 500, value % 12 + 1, 1)


def _month_end(value: date) -> date:
    next_month = (
        date(value.year + 1, 1, 1) if value.month == 12 else date(value.year, value.month + 1, 1)
    )
    return date.fromordinal(next_month.toordinal() - 1)


def _rule_request(
    bank_id: str,
    product_id: str,
    period: date,
    *,
    source: str = "case_owner",
    level: int | None = None,
    fixed_amount: str = "100.00",
) -> dict[str, object]:
    return {
        "bank_id": bank_id,
        "product_id": product_id,
        "eligibility_milestone": "booked",
        "effective_from": period.isoformat(),
        "effective_to": _month_end(period).isoformat(),
        "payout_mode": "percentage_split",
        "calculation_method": "fixed",
        "fixed_amount": fixed_amount,
        "percentage_rate": None,
        "flat_amount": None,
        "slabs": [],
        "recipients": [
            {
                "role_code": "case_owner" if source == "case_owner" else f"manager_{level}",
                "role_name": "Case Owner" if source == "case_owner" else f"Manager {level}",
                "recipient_source": source,
                "hierarchy_level": level,
                "sort_order": 0,
                "split_percent": "100",
                "calculation_method": None,
                "fixed_amount": None,
                "percentage_rate": None,
                "flat_amount": None,
                "slabs": [],
            }
        ],
    }


async def _activate_rule(
    client: AsyncClient,
    bank_id: str,
    product_id: str,
    period: date,
    *,
    source: str = "case_owner",
    level: int | None = None,
    fixed_amount: str = "100.00",
) -> dict:
    created = await client.post(
        "/api/v1/finance/commission-rules",
        json=_rule_request(
            bank_id,
            product_id,
            period,
            source=source,
            level=level,
            fixed_amount=fixed_amount,
        ),
    )
    assert created.status_code == 200, created.text
    activated = await client.post(
        f"/api/v1/finance/commission-rules/{created.json()['id']}/activate"
    )
    assert activated.status_code == 200, activated.text
    return activated.json()


async def _booked_application(
    client: AsyncClient,
    *,
    bank_id: str,
    product_id: str,
    owner_id: str,
    event_at: datetime,
    booked_amount: str = "1000.00",
) -> dict:
    created = await _create_app(
        client,
        customer_id=(await _customer(client, "Finance"))["id"],
        bank_id=bank_id,
        product_id=product_id,
        case_owner_id=owner_id,
        requested_amount="999999.99",
    )
    async with app.state.session_factory() as session:
        row = await session.get(Application, UUID(created["id"]))
        assert row is not None
        row.approved_amount = Decimal("888888.88")
        row.booked_amount = Decimal(booked_amount)
        row.booked_at = event_at
        await session.commit()
    return created


_DUMMY_ID = "00000000-0000-0000-0000-000000000001"
_DUMMY_MONTH = "2200-01-01"
_RULE_BODY = _rule_request(_DUMMY_ID, _DUMMY_ID, date(2200, 1, 1))
_PLAN_BODY = {
    "name": "Security plan",
    "effective_from": _DUMMY_MONTH,
    "effective_to": "2200-01-31",
    "slabs": [
        {
            "minimum_production": "0",
            "maximum_production": None,
            "payout_amount": "10",
            "sort_order": 0,
        }
    ],
}

_ROUTE_MATRIX = (
    ("GET", "/api/v1/finance/options", None, "Finance.ViewCommissionRules"),
    ("GET", "/api/v1/finance/commission-rules", None, "Finance.ViewCommissionRules"),
    ("POST", "/api/v1/finance/commission-rules", _RULE_BODY, "Finance.ManageCommissionRules"),
    (
        "GET",
        f"/api/v1/finance/commission-rules/{_DUMMY_ID}",
        None,
        "Finance.ViewCommissionRules",
    ),
    (
        "POST",
        f"/api/v1/finance/commission-rules/{_DUMMY_ID}/activate",
        None,
        "Finance.ManageCommissionRules",
    ),
    (
        "POST",
        f"/api/v1/finance/commission-rules/{_DUMMY_ID}/deactivate",
        None,
        "Finance.ManageCommissionRules",
    ),
    ("GET", "/api/v1/finance/incentive-plans", None, "Finance.ViewCommissionRules"),
    ("POST", "/api/v1/finance/incentive-plans", _PLAN_BODY, "Finance.ManageCommissionRules"),
    (
        "POST",
        f"/api/v1/finance/incentive-plans/{_DUMMY_ID}/activate",
        None,
        "Finance.ManageCommissionRules",
    ),
    (
        "POST",
        f"/api/v1/finance/incentive-plans/{_DUMMY_ID}/deactivate",
        None,
        "Finance.ManageCommissionRules",
    ),
    ("GET", "/api/v1/finance/periods", None, "Finance.View"),
    ("GET", f"/api/v1/finance/periods/{_DUMMY_ID}", None, "Finance.View"),
    (
        "POST",
        f"/api/v1/finance/periods/{_DUMMY_MONTH}/generate",
        None,
        "Finance.GeneratePayout",
    ),
    (
        "POST",
        f"/api/v1/finance/periods/{_DUMMY_MONTH}/review",
        None,
        "Finance.Review",
    ),
    (
        "POST",
        f"/api/v1/finance/periods/{_DUMMY_MONTH}/finalize",
        None,
        "Finance.Finalize",
    ),
    (
        "POST",
        f"/api/v1/finance/periods/{_DUMMY_MONTH}/reopen",
        {"reason": "Security test"},
        "Finance.ReopenPeriod",
    ),
    (
        "POST",
        f"/api/v1/finance/periods/{_DUMMY_MONTH}/adjustments",
        {
            "application_id": _DUMMY_ID,
            "recipient_id": _DUMMY_ID,
            "amount": "1.00",
            "reason": "Security test",
        },
        "Finance.EditAdjustment",
    ),
    (
        "POST",
        f"/api/v1/finance/periods/{_DUMMY_MONTH}/clawbacks",
        {"original_component_id": _DUMMY_ID, "amount": "1.00", "reason": "Security test"},
        "Finance.EditAdjustment",
    ),
    (
        "GET",
        f"/api/v1/finance/statements?period_month={_DUMMY_MONTH}",
        None,
        "Finance.View",
    ),
    (
        "GET",
        f"/api/v1/finance/payouts/{_DUMMY_ID}/components",
        None,
        "Finance.View",
    ),
    (
        "POST",
        "/api/v1/finance/export",
        {"format": "xlsx", "period_month": _DUMMY_MONTH, "recipient_id": None},
        "Finance.View",
    ),
)


def test_finance_security_matrix_covers_every_registered_route() -> None:
    registered = {
        (method, f"/api/v1{route.path}")
        for route in finance_router.routes
        for method in route.methods
        if method in {"GET", "POST"}
    }
    expected = {
        ("GET", "/api/v1/finance/options"),
        ("GET", "/api/v1/finance/commission-rules"),
        ("POST", "/api/v1/finance/commission-rules"),
        ("GET", "/api/v1/finance/commission-rules/{rule_id}"),
        ("POST", "/api/v1/finance/commission-rules/{rule_id}/activate"),
        ("POST", "/api/v1/finance/commission-rules/{rule_id}/deactivate"),
        ("GET", "/api/v1/finance/incentive-plans"),
        ("POST", "/api/v1/finance/incentive-plans"),
        ("POST", "/api/v1/finance/incentive-plans/{plan_id}/activate"),
        ("POST", "/api/v1/finance/incentive-plans/{plan_id}/deactivate"),
        ("GET", "/api/v1/finance/periods"),
        ("GET", "/api/v1/finance/periods/{period_id}"),
        ("POST", "/api/v1/finance/periods/{period_month}/generate"),
        ("POST", "/api/v1/finance/periods/{period_month}/review"),
        ("POST", "/api/v1/finance/periods/{period_month}/finalize"),
        ("POST", "/api/v1/finance/periods/{period_month}/reopen"),
        ("POST", "/api/v1/finance/periods/{period_month}/adjustments"),
        ("POST", "/api/v1/finance/periods/{period_month}/clawbacks"),
        ("GET", "/api/v1/finance/statements"),
        ("GET", "/api/v1/finance/payouts/{payout_id}/components"),
        ("POST", "/api/v1/finance/export"),
    }
    assert registered == expected


@pytest.mark.asyncio
async def test_every_finance_route_requires_authentication(client: AsyncClient) -> None:
    for method, path, body, _permission in _ROUTE_MATRIX:
        response = await client.request(method, path, json=body)
        assert response.status_code == 401, (method, path, response.text)
        assert response.json()["error"]["code"] == "UNAUTHENTICATED"


@pytest.mark.asyncio
async def test_every_finance_route_enforces_its_permission(client: AsyncClient) -> None:
    authed, _owner = await owner_client(client)
    denied_user = await _reporting_user(
        authed,
        scope="company",
        permissions=["Users.View"],
    )
    async with await spawned_client() as denied:
        await authenticate(denied, denied_user["email"], "UserPass1!")
        for method, path, body, permission in _ROUTE_MATRIX:
            response = await denied.request(method, path, json=body)
            assert response.status_code == 403, (method, path, response.text)
            error = response.json()["error"]
            assert error["code"] == "FORBIDDEN"
            assert error["details"] == [{"permission": permission}]


@pytest.mark.asyncio
async def test_every_state_changing_finance_route_preserves_csrf(client: AsyncClient) -> None:
    authed, _owner = await owner_client(client)
    csrf = authed.headers.pop("X-CSRF-Token")
    try:
        for method, path, body, _permission in _ROUTE_MATRIX:
            if method != "POST":
                continue
            response = await authed.request(method, path, json=body)
            assert response.status_code == 403, (method, path, response.text)
            assert response.json()["error"]["code"] == "CSRF_INVALID"
    finally:
        authed.headers["X-CSRF-Token"] = csrf


@pytest.mark.asyncio
async def test_finance_request_models_reject_mass_assignment(client: AsyncClient) -> None:
    authed, _owner = await owner_client(client)
    nested_rule = {
        **_RULE_BODY,
        "recipients": [{**_RULE_BODY["recipients"][0], "user_id": _DUMMY_ID}],
    }
    attempts = (
        ("/api/v1/finance/commission-rules", {**_RULE_BODY, "status": "active"}),
        ("/api/v1/finance/commission-rules", nested_rule),
        ("/api/v1/finance/incentive-plans", {**_PLAN_BODY, "activated_by_id": _DUMMY_ID}),
        (
            f"/api/v1/finance/periods/{_DUMMY_MONTH}/adjustments",
            {
                "application_id": _DUMMY_ID,
                "recipient_id": _DUMMY_ID,
                "amount": "1.00",
                "reason": "test",
                "actor_id": _DUMMY_ID,
            },
        ),
        (
            f"/api/v1/finance/periods/{_DUMMY_MONTH}/clawbacks",
            {
                "original_component_id": _DUMMY_ID,
                "amount": "1.00",
                "reason": "test",
                "recipient_id": _DUMMY_ID,
            },
        ),
        (
            f"/api/v1/finance/periods/{_DUMMY_MONTH}/reopen",
            {"reason": "test", "status": "draft"},
        ),
        (
            "/api/v1/finance/export",
            {
                "format": "xlsx",
                "period_month": _DUMMY_MONTH,
                "recipient_id": None,
                "reporting_scope": "company",
            },
        ),
    )
    for path, body in attempts:
        response = await authed.post(path, json=body)
        assert response.status_code == 422, (path, response.text)


@pytest.mark.asyncio
async def test_generation_fails_closed_without_partial_persistence(client: AsyncClient) -> None:
    authed, _owner = await owner_client(client)
    dib, _eib, pf, _cc = await _catalog(authed)
    period = _unique_month()
    event_at = datetime(period.year, period.month, 10, 12, tzinfo=UTC)
    manager = await _reporting_user(
        authed,
        scope="company",
        permissions=["Finance.View"],
        can_be_reporting_manager=True,
    )
    valid_owner = await _reporting_user(
        authed,
        scope="own",
        permissions=["Finance.View"],
        manager_id=manager["id"],
    )
    invalid_owner = await _reporting_user(
        authed,
        scope="own",
        permissions=["Finance.View"],
    )
    await _booked_application(
        authed,
        bank_id=dib["id"],
        product_id=pf["id"],
        owner_id=valid_owner["id"],
        event_at=event_at,
    )
    invalid_application = await _booked_application(
        authed,
        bank_id=dib["id"],
        product_id=pf["id"],
        owner_id=invalid_owner["id"],
        event_at=event_at,
    )
    await _activate_rule(
        authed,
        dib["id"],
        pf["id"],
        period,
        source="reporting_manager",
        level=1,
    )
    async with app.state.session_factory() as session:
        component_before = await session.scalar(select(func.count()).select_from(FinanceComponent))
        payout_before = await session.scalar(select(func.count()).select_from(FinancePayout))
        transition_before = await session.scalar(
            select(func.count()).select_from(FinancePeriodTransition)
        )
        audit_before = await session.scalar(
            select(func.count())
            .select_from(AuditEvent)
            .where(AuditEvent.action == "finance.period.generate")
        )
    response = await authed.post(f"/api/v1/finance/periods/{period.isoformat()}/generate")
    assert response.status_code == 422, response.text
    error = response.json()["error"]
    assert error["code"] == "FINANCE_RECIPIENT_UNRESOLVED"
    assert error["details"] == [
        {
            "applicationId": invalid_application["id"],
            "applicationCode": invalid_application["applicationCode"],
            "source": "reporting_manager",
            "hierarchyLevel": 1,
            "eligibilityEventAt": event_at.isoformat(),
        }
    ]
    async with app.state.session_factory() as session:
        payout_period = (
            await session.execute(
                select(FinancePayoutPeriod).where(FinancePayoutPeriod.period_month == period)
            )
        ).scalar_one_or_none()
        assert payout_period is None
        assert (
            await session.scalar(select(func.count()).select_from(FinanceComponent))
            == component_before
        )
        assert (
            await session.scalar(select(func.count()).select_from(FinancePayout)) == payout_before
        )
        assert (
            await session.scalar(select(func.count()).select_from(FinancePeriodTransition))
            == transition_before
        )
        audit_after = await session.scalar(
            select(func.count())
            .select_from(AuditEvent)
            .where(AuditEvent.action == "finance.period.generate")
        )
        assert audit_after == audit_before


@pytest.mark.asyncio
async def test_server_scope_blocks_idor_tampering_and_export_leakage(
    client: AsyncClient,
) -> None:
    authed, _owner = await owner_client(client)
    dib, _eib, pf, _cc = await _catalog(authed)
    offices = {item["code"]: item for item in (await authed.get("/api/v1/offices")).json()["items"]}
    period = _unique_month()
    event_at = datetime(period.year, period.month, 11, 12, tzinfo=UTC)
    visible_user = await _reporting_user(
        authed,
        scope="office",
        permissions=["Finance.View", "Finance.EditAdjustment"],
        office_id=offices["DXB"]["id"],
    )
    hidden_user = await _reporting_user(
        authed,
        scope="own",
        permissions=["Finance.View"],
        office_id=offices["AUH"]["id"],
    )
    visible_application = await _booked_application(
        authed,
        bank_id=dib["id"],
        product_id=pf["id"],
        owner_id=visible_user["id"],
        event_at=event_at,
    )
    hidden_application = await _booked_application(
        authed,
        bank_id=dib["id"],
        product_id=pf["id"],
        owner_id=hidden_user["id"],
        event_at=event_at,
    )
    await _activate_rule(authed, dib["id"], pf["id"], period)
    generated = await authed.post(f"/api/v1/finance/periods/{period.isoformat()}/generate")
    assert generated.status_code == 200, generated.text
    owner_period = generated.json()
    payouts = {row["recipientId"]: row for row in owner_period["payouts"]}
    visible_payout = payouts[visible_user["id"]]
    hidden_payout = payouts[hidden_user["id"]]
    owner_hidden_drill = await authed.get(
        f"/api/v1/finance/payouts/{hidden_payout['id']}/components"
    )
    hidden_component = owner_hidden_drill.json()["items"][0]

    async with await spawned_client() as scoped:
        await authenticate(scoped, visible_user["email"], "UserPass1!")
        period_response = await scoped.get(
            f"/api/v1/finance/periods/{owner_period['id']}?reporting_scope=company"
        )
        assert period_response.status_code == 200, period_response.text
        assert {row["recipientId"] for row in period_response.json()["payouts"]} == {
            visible_user["id"]
        }
        tampered = await scoped.get(
            "/api/v1/finance/statements",
            params={
                "period_month": period.isoformat(),
                "recipient_id": hidden_user["id"],
                "reporting_scope": "company",
            },
        )
        assert tampered.status_code == 404
        assert tampered.json()["error"]["code"] == "FINANCE_STATEMENT_NOT_FOUND"
        hidden_drill = await scoped.get(f"/api/v1/finance/payouts/{hidden_payout['id']}/components")
        assert hidden_drill.status_code == 404
        hidden_adjustment = await scoped.post(
            f"/api/v1/finance/periods/{period.isoformat()}/adjustments",
            json={
                "application_id": hidden_application["id"],
                "recipient_id": hidden_user["id"],
                "amount": "1.00",
                "reason": "IDOR attempt",
            },
        )
        assert hidden_adjustment.status_code == 404
        assert hidden_adjustment.json()["error"]["code"] == "APPLICATION_NOT_FOUND"
        tampered_recipient = await scoped.post(
            f"/api/v1/finance/periods/{period.isoformat()}/adjustments",
            json={
                "application_id": visible_application["id"],
                "recipient_id": hidden_user["id"],
                "amount": "1.00",
                "reason": "Recipient tampering attempt",
            },
        )
        assert tampered_recipient.status_code == 422
        assert tampered_recipient.json()["error"]["code"] == "FINANCE_ATTRIBUTION_MISMATCH"
        hidden_clawback = await scoped.post(
            f"/api/v1/finance/periods/{period.isoformat()}/clawbacks",
            json={
                "original_component_id": hidden_component["id"],
                "amount": "1.00",
                "reason": "Component IDOR attempt",
            },
        )
        assert hidden_clawback.status_code == 404
        assert hidden_clawback.json()["error"]["code"] == "APPLICATION_NOT_FOUND"
        visible_drill = await scoped.get(
            f"/api/v1/finance/payouts/{visible_payout['id']}/components"
        )
        assert visible_drill.status_code == 200, visible_drill.text
        assert {item["applicationId"] for item in visible_drill.json()["items"]} == {
            visible_application["id"]
        }
        assert hidden_application["id"] not in visible_drill.text
        for item in visible_drill.json()["items"]:
            assert "attributionSnapshot" not in item
            assert "calculationEvidence" not in item
            assert "actorId" not in item

        exported = await scoped.post(
            "/api/v1/finance/export",
            json={"format": "xlsx", "period_month": period.isoformat(), "recipient_id": None},
        )
        assert exported.status_code == 200, exported.text
        workbook = load_workbook(BytesIO(exported.content), read_only=True, data_only=True)
        rows = list(workbook["Finance Statement"].iter_rows(values_only=True))
        flattened = "\n".join("|".join(str(value or "") for value in row) for row in rows)
        assert visible_user["userCode"] in flattened
        assert hidden_user["userCode"] not in flattened


@pytest.mark.asyncio
async def test_visible_payout_never_exposes_an_out_of_scope_application(
    client: AsyncClient,
) -> None:
    authed, _owner = await owner_client(client)
    dib, _eib, pf, _cc = await _catalog(authed)
    offices = {item["code"]: item for item in (await authed.get("/api/v1/offices")).json()["items"]}
    period = _unique_month()
    event_at = datetime(period.year, period.month, 13, 12, tzinfo=UTC)
    manager = await _reporting_user(
        authed,
        scope="own",
        permissions=["Finance.View"],
        office_id=offices["DXB"]["id"],
        can_be_reporting_manager=True,
    )
    case_owner = await _reporting_user(
        authed,
        scope="own",
        permissions=["Finance.View"],
        office_id=offices["AUH"]["id"],
        manager_id=manager["id"],
    )
    hidden_application = await _booked_application(
        authed,
        bank_id=dib["id"],
        product_id=pf["id"],
        owner_id=case_owner["id"],
        event_at=event_at,
    )
    await _activate_rule(
        authed,
        dib["id"],
        pf["id"],
        period,
        source="reporting_manager",
        level=1,
    )
    generated = await authed.post(f"/api/v1/finance/periods/{period.isoformat()}/generate")
    assert generated.status_code == 200, generated.text
    payout = next(row for row in generated.json()["payouts"] if row["recipientId"] == manager["id"])

    async with await spawned_client() as scoped:
        await authenticate(scoped, manager["email"], "UserPass1!")
        statement = await scoped.get(
            "/api/v1/finance/statements", params={"period_month": period.isoformat()}
        )
        assert statement.status_code == 200, statement.text
        assert statement.json()["items"][0]["recipientId"] == manager["id"]
        assert statement.json()["items"][0]["eligibleCases"] == 0
        assert statement.json()["items"][0]["eligibleValue"] == "0.00"
        drill = await scoped.get(f"/api/v1/finance/payouts/{payout['id']}/components")
        assert drill.status_code == 200, drill.text
        assert drill.json()["items"] == []
        assert drill.json()["total"] == 0
        assert hidden_application["id"] not in drill.text


@pytest.mark.asyncio
async def test_audit_and_finalized_period_immutability_are_enforced(client: AsyncClient) -> None:
    authed, owner = await owner_client(client)
    dib, _eib, pf, _cc = await _catalog(authed)
    period = _unique_month()
    event_at = datetime(period.year, period.month, 12, 12, tzinfo=UTC)
    application = await _booked_application(
        authed,
        bank_id=dib["id"],
        product_id=pf["id"],
        owner_id=owner["id"],
        event_at=event_at,
    )
    rule = await _activate_rule(authed, dib["id"], pf["id"], period)
    generated = await authed.post(f"/api/v1/finance/periods/{period.isoformat()}/generate")
    assert generated.status_code == 200, generated.text
    deactivated_rule = await authed.post(
        f"/api/v1/finance/commission-rules/{rule['id']}/deactivate"
    )
    assert deactivated_rule.status_code == 200, deactivated_rule.text
    plan_created = await authed.post(
        "/api/v1/finance/incentive-plans",
        json={
            "name": f"Audit plan {unique_tag()}",
            "effective_from": period.isoformat(),
            "effective_to": _month_end(period).isoformat(),
            "slabs": [
                {
                    "minimum_production": "0.00",
                    "maximum_production": None,
                    "payout_amount": "10.00",
                    "sort_order": 0,
                }
            ],
        },
    )
    assert plan_created.status_code == 200, plan_created.text
    plan_id = plan_created.json()["id"]
    plan_activated = await authed.post(f"/api/v1/finance/incentive-plans/{plan_id}/activate")
    assert plan_activated.status_code == 200, plan_activated.text
    plan_deactivated = await authed.post(f"/api/v1/finance/incentive-plans/{plan_id}/deactivate")
    assert plan_deactivated.status_code == 200, plan_deactivated.text
    payout = generated.json()["payouts"][0]
    drill = await authed.get(f"/api/v1/finance/payouts/{payout['id']}/components")
    original = next(item for item in drill.json()["items"] if item["componentType"] == "commission")
    blank_adjustment = await authed.post(
        f"/api/v1/finance/periods/{period.isoformat()}/adjustments",
        json={
            "application_id": application["id"],
            "recipient_id": owner["id"],
            "amount": "1.00",
            "reason": "   ",
        },
    )
    assert blank_adjustment.status_code == 422
    assert blank_adjustment.json()["error"]["code"] == "REASON_REQUIRED"
    blank_clawback = await authed.post(
        f"/api/v1/finance/periods/{period.isoformat()}/clawbacks",
        json={"original_component_id": original["id"], "amount": "1.00", "reason": "   "},
    )
    assert blank_clawback.status_code == 422
    assert blank_clawback.json()["error"]["code"] == "REASON_REQUIRED"
    adjustment = await authed.post(
        f"/api/v1/finance/periods/{period.isoformat()}/adjustments",
        json={
            "application_id": application["id"],
            "recipient_id": owner["id"],
            "amount": "12.345",
            "reason": "Authorized positive correction",
        },
    )
    assert adjustment.status_code == 200, adjustment.text
    assert adjustment.json()["amount"] == "12.35"
    clawback = await authed.post(
        f"/api/v1/finance/periods/{period.isoformat()}/clawbacks",
        json={
            "original_component_id": original["id"],
            "amount": "5.555",
            "reason": "Authorized recovery",
        },
    )
    assert clawback.status_code == 200, clawback.text
    assert clawback.json()["amount"] == "-5.56"

    first_review, second_review = await asyncio.gather(
        authed.post(f"/api/v1/finance/periods/{period.isoformat()}/review"),
        authed.post(f"/api/v1/finance/periods/{period.isoformat()}/review"),
    )
    assert sorted((first_review.status_code, second_review.status_code)) == [200, 409]
    finalized = await authed.post(f"/api/v1/finance/periods/{period.isoformat()}/finalize")
    assert finalized.status_code == 200, finalized.text
    before = await authed.get(f"/api/v1/finance/payouts/{payout['id']}/components")
    before_ids = [item["id"] for item in before.json()["items"]]
    locked_adjustment = await authed.post(
        f"/api/v1/finance/periods/{period.isoformat()}/adjustments",
        json={
            "application_id": application["id"],
            "recipient_id": owner["id"],
            "amount": "1.00",
            "reason": "Must be rejected",
        },
    )
    assert locked_adjustment.status_code == 409
    assert locked_adjustment.json()["error"]["code"] == "FINANCE_PERIOD_LOCKED"
    locked_clawback = await authed.post(
        f"/api/v1/finance/periods/{period.isoformat()}/clawbacks",
        json={
            "original_component_id": original["id"],
            "amount": "1.00",
            "reason": "Must be rejected",
        },
    )
    assert locked_clawback.status_code == 409
    assert locked_clawback.json()["error"]["code"] == "FINANCE_PERIOD_LOCKED"
    after = await authed.get(f"/api/v1/finance/payouts/{payout['id']}/components")
    assert [item["id"] for item in after.json()["items"]] == before_ids
    assert (await authed.delete(f"/api/v1/finance/periods/{period.isoformat()}")).status_code == 405

    blank_reopen = await authed.post(
        f"/api/v1/finance/periods/{period.isoformat()}/reopen", json={"reason": "   "}
    )
    assert blank_reopen.status_code == 422
    assert blank_reopen.json()["error"]["code"] == "REASON_REQUIRED"
    reopened = await authed.post(
        f"/api/v1/finance/periods/{period.isoformat()}/reopen",
        json={"reason": "Approved correction cycle"},
    )
    assert reopened.status_code == 200, reopened.text
    assert reopened.json()["status"] == "review"
    assert [
        item["id"]
        for item in (await authed.get(f"/api/v1/finance/payouts/{payout['id']}/components")).json()[
            "items"
        ]
    ] == before_ids

    exported = await authed.post(
        "/api/v1/finance/export",
        json={"format": "print", "period_month": period.isoformat(), "recipient_id": owner["id"]},
    )
    assert exported.status_code == 200
    async with app.state.session_factory() as session:
        audits = list(
            (
                await session.execute(
                    select(AuditEvent).where(
                        AuditEvent.entity_id.in_(
                            (
                                rule["id"],
                                plan_id,
                                generated.json()["id"],
                                adjustment.json()["id"],
                                clawback.json()["id"],
                                period.isoformat(),
                            )
                        )
                    )
                )
            ).scalars()
        )
        pairs = {(row.action, row.entity_id) for row in audits}
        assert ("finance.rule.create", rule["id"]) in pairs
        assert ("finance.rule.activate", rule["id"]) in pairs
        assert ("finance.rule.deactivate", rule["id"]) in pairs
        assert ("finance.incentive_plan.create", plan_id) in pairs
        assert ("finance.incentive_plan.activate", plan_id) in pairs
        assert ("finance.incentive_plan.deactivate", plan_id) in pairs
        assert ("finance.period.generate", generated.json()["id"]) in pairs
        assert ("finance.adjustment.create", adjustment.json()["id"]) in pairs
        assert ("finance.clawback.create", clawback.json()["id"]) in pairs
        assert ("finance.period.review", generated.json()["id"]) in pairs
        assert ("finance.period.finalize", generated.json()["id"]) in pairs
        assert ("finance.period.reopen", generated.json()["id"]) in pairs
        assert ("finance.statement.export", period.isoformat()) in pairs
        component = await session.get(FinanceComponent, UUID(original["id"]))
        assert component is not None
        component.amount = Decimal("999.00")
        with pytest.raises(RuntimeError, match="immutable"):
            await session.flush()
        await session.rollback()

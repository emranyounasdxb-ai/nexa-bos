from __future__ import annotations

from io import BytesIO
from uuid import UUID

import pytest
from helpers import (
    authenticate,
    create_activated_user,
    office_id,
    owner_client,
    spawned_client,
    unique_tag,
)
from httpx import AsyncClient
from nexa_bos_api.assets.api import router as assets_router
from nexa_bos_api.identity.models import AuditEvent
from nexa_bos_api.main import app
from openpyxl import load_workbook
from sqlalchemy import func, select
from test_assets import _category, _create_pc, _employee, _pc_payload

_DUMMY_ID = "00000000-0000-0000-0000-000000000001"

_ROUTE_MATRIX = (
    ("GET", "/api/v1/assets/options", None, "Assets.View"),
    ("GET", "/api/v1/assets/categories", None, "Assets.View"),
    (
        "POST",
        "/api/v1/assets/categories",
        {"code": "SEC", "name": "Security", "fields": []},
        "Assets.ManageMaster",
    ),
    (
        "PATCH",
        f"/api/v1/assets/categories/{_DUMMY_ID}",
        {"name": "Security"},
        "Assets.ManageMaster",
    ),
    (
        "POST",
        f"/api/v1/assets/categories/{_DUMMY_ID}/activate",
        None,
        "Assets.ManageMaster",
    ),
    (
        "POST",
        f"/api/v1/assets/categories/{_DUMMY_ID}/deactivate",
        None,
        "Assets.ManageMaster",
    ),
    ("GET", "/api/v1/assets/reports/asset_register", None, "Assets.View"),
    (
        "POST",
        "/api/v1/assets/reports/export",
        {"format": "xlsx", "report": "asset_register"},
        "Assets.View",
    ),
    ("GET", "/api/v1/assets/audit", None, "Assets.ViewAudit"),
    ("GET", f"/api/v1/assets/employees/{_DUMMY_ID}", None, "Assets.View"),
    ("GET", "/api/v1/assets", None, "Assets.View"),
    (
        "POST",
        "/api/v1/assets",
        {
            "category_id": _DUMMY_ID,
            "office_id": _DUMMY_ID,
            "condition": "Good",
            "attributes": {},
        },
        "Assets.ManageStock",
    ),
    ("GET", f"/api/v1/assets/{_DUMMY_ID}", None, "Assets.View"),
    (
        "PATCH",
        f"/api/v1/assets/{_DUMMY_ID}",
        {"brand": "Security"},
        "Assets.ManageMaster",
    ),
    (
        "POST",
        f"/api/v1/assets/{_DUMMY_ID}/identifiers",
        {"serial_number": "SEC", "reason": "Security"},
        "Assets.ManageMaster",
    ),
    (
        "POST",
        f"/api/v1/assets/{_DUMMY_ID}/condition",
        {"condition": "Good", "reason": "Security"},
        "Assets.ManageStock",
    ),
    (
        "POST",
        f"/api/v1/assets/{_DUMMY_ID}/allocate",
        {
            "employee_id": _DUMMY_ID,
            "issue_date": "2026-08-30",
            "condition_at_issue": "Good",
        },
        "Assets.Allocate",
    ),
    (
        "POST",
        f"/api/v1/assets/{_DUMMY_ID}/return",
        {"return_date": "2026-08-30", "return_condition": "Good"},
        "Assets.Return",
    ),
    (
        "POST",
        f"/api/v1/assets/{_DUMMY_ID}/transfer/employee",
        {
            "employee_id": _DUMMY_ID,
            "transfer_date": "2026-08-30",
            "condition": "Good",
        },
        "Assets.Transfer",
    ),
    (
        "POST",
        f"/api/v1/assets/{_DUMMY_ID}/transfer/office",
        {"office_id": _DUMMY_ID, "transfer_date": "2026-08-30"},
        "Assets.Transfer",
    ),
    (
        "POST",
        f"/api/v1/assets/{_DUMMY_ID}/status",
        {"status": "Lost", "reason": "Security"},
        "Assets.ManageStatus",
    ),
    ("GET", f"/api/v1/assets/{_DUMMY_ID}/history", None, "Assets.ViewAudit"),
)


async def _type_with(
    client: AsyncClient,
    permissions: list[str],
    *,
    scope: str,
) -> str:
    tag = unique_tag()[:8].upper()
    code = f"S{tag}"
    created = await client.post(
        "/api/v1/user-types",
        json={"name": f"Asset Security {tag}", "code": code},
    )
    assert created.status_code == 200, created.text
    type_id = created.json()["id"]
    assert (await client.post(f"/api/v1/user-types/{type_id}/activate")).status_code == 200
    assigned = await client.put(
        f"/api/v1/user-types/{type_id}/permissions",
        json={"permissions": permissions},
    )
    assert assigned.status_code == 200, assigned.text
    scoped = await client.put(
        f"/api/v1/user-types/{type_id}/scope",
        json={"visibility_scope": scope},
    )
    assert scoped.status_code == 200, scoped.text
    return code


def test_asset_security_matrix_covers_every_registered_route() -> None:
    registered = {
        (method, f"/api/v1{route.path}")
        for route in assets_router.routes
        for method in route.methods
        if method in {"GET", "POST", "PATCH", "DELETE"}
    }
    expected = {
        ("GET", "/api/v1/assets/options"),
        ("GET", "/api/v1/assets/categories"),
        ("POST", "/api/v1/assets/categories"),
        ("PATCH", "/api/v1/assets/categories/{category_id}"),
        ("POST", "/api/v1/assets/categories/{category_id}/activate"),
        ("POST", "/api/v1/assets/categories/{category_id}/deactivate"),
        ("GET", "/api/v1/assets/reports/{report}"),
        ("POST", "/api/v1/assets/reports/export"),
        ("GET", "/api/v1/assets/audit"),
        ("GET", "/api/v1/assets/employees/{employee_id}"),
        ("GET", "/api/v1/assets"),
        ("POST", "/api/v1/assets"),
        ("GET", "/api/v1/assets/{asset_id}"),
        ("PATCH", "/api/v1/assets/{asset_id}"),
        ("POST", "/api/v1/assets/{asset_id}/identifiers"),
        ("POST", "/api/v1/assets/{asset_id}/condition"),
        ("POST", "/api/v1/assets/{asset_id}/allocate"),
        ("POST", "/api/v1/assets/{asset_id}/return"),
        ("POST", "/api/v1/assets/{asset_id}/transfer/employee"),
        ("POST", "/api/v1/assets/{asset_id}/transfer/office"),
        ("POST", "/api/v1/assets/{asset_id}/status"),
        ("GET", "/api/v1/assets/{asset_id}/history"),
    }
    assert registered == expected


@pytest.mark.asyncio
async def test_every_asset_route_requires_authentication(client: AsyncClient) -> None:
    for method, path, body, _permission in _ROUTE_MATRIX:
        response = await client.request(method, path, json=body)
        assert response.status_code == 401, (method, path, response.text)
        assert response.json()["error"]["code"] == "UNAUTHENTICATED"


@pytest.mark.asyncio
async def test_every_asset_route_enforces_exact_permission(client: AsyncClient) -> None:
    owner, _ = await owner_client(client)
    denied_type = await _type_with(owner, ["Users.View"], scope="company")
    denied_user = await create_activated_user(owner, user_type_code=denied_type)
    async with await spawned_client() as denied:
        await authenticate(denied, denied_user["email"], "UserPass1!")
        for method, path, body, permission in _ROUTE_MATRIX:
            response = await denied.request(method, path, json=body)
            assert response.status_code == 403, (method, path, response.text)
            error = response.json()["error"]
            assert error["code"] == "FORBIDDEN"
            assert error["details"] == [{"permission": permission}]


@pytest.mark.asyncio
async def test_every_asset_mutation_preserves_csrf(client: AsyncClient) -> None:
    owner, _ = await owner_client(client)
    csrf = owner.headers.pop("X-CSRF-Token")
    try:
        probes = [row for row in _ROUTE_MATRIX if row[0] in {"POST", "PATCH"}]
        for method, path, body, _permission in probes:
            response = await owner.request(method, path, json=body)
            assert response.status_code == 403, (method, path, response.text)
            assert response.json()["error"]["code"] == "CSRF_INVALID"
    finally:
        owner.headers["X-CSRF-Token"] = csrf


@pytest.mark.asyncio
async def test_damaged_return_requires_status_authority_and_reason(
    client: AsyncClient,
) -> None:
    owner, _ = await owner_client(client)
    dxb = await office_id(owner, "DXB")
    employee = await _employee(owner, dxb)
    asset = await _create_pc(owner, dxb)
    allocated = await owner.post(
        f"/api/v1/assets/{asset['id']}/allocate",
        json={
            "employee_id": employee["id"],
            "issue_date": "2026-08-01",
            "condition_at_issue": "Good",
        },
    )
    assert allocated.status_code == 200, allocated.text

    return_type = await _type_with(
        owner,
        ["Assets.View", "Assets.Return"],
        scope="company",
    )
    return_user = await create_activated_user(owner, user_type_code=return_type)
    async with await spawned_client() as restricted:
        await authenticate(restricted, return_user["email"], "UserPass1!")
        denied = await restricted.post(
            f"/api/v1/assets/{asset['id']}/return",
            json={
                "return_date": "2026-08-20",
                "return_condition": "Damaged",
                "remarks": "Screen cracked",
            },
        )
        assert denied.status_code == 403, denied.text
        assert denied.json()["error"]["details"] == [{"permission": "Assets.ManageStatus"}]

    missing_reason = await owner.post(
        f"/api/v1/assets/{asset['id']}/return",
        json={"return_date": "2026-08-20", "return_condition": "Damaged"},
    )
    assert missing_reason.status_code == 422
    assert missing_reason.json()["error"]["code"] == "ASSET_STATUS_REASON_REQUIRED"
    unchanged = await owner.get(f"/api/v1/assets/{asset['id']}")
    assert unchanged.json()["status"] == "Allocated"
    assert unchanged.json()["currentAllocation"]["employeeId"] == employee["id"]

    returned = await owner.post(
        f"/api/v1/assets/{asset['id']}/return",
        json={
            "return_date": "2026-08-20",
            "return_condition": "Damaged",
            "remarks": "Screen cracked",
        },
    )
    assert returned.status_code == 200, returned.text
    assert returned.json()["status"] == "Damaged"
    history = await owner.get(f"/api/v1/assets/{asset['id']}/history")
    event = next(row for row in history.json()["events"] if row["action"] == "asset.return")
    assert event["reason"] == "Screen cracked"


@pytest.mark.parametrize("controlled_status", ["Lost", "Damaged", "Under Repair"])
@pytest.mark.parametrize(
    ("operation", "permission"),
    [
        ("return", "Assets.Return"),
        ("employee_transfer", "Assets.Transfer"),
    ],
)
@pytest.mark.asyncio
async def test_h1_controlled_status_blocks_custody_operation_until_authorized_correction(
    client: AsyncClient,
    controlled_status: str,
    operation: str,
    permission: str,
) -> None:
    owner, _ = await owner_client(client)
    dxb = await office_id(owner, "DXB")
    current_employee = await _employee(owner, dxb)
    next_employee = await _employee(owner, dxb)
    asset = await _create_pc(owner, dxb)
    allocated = await owner.post(
        f"/api/v1/assets/{asset['id']}/allocate",
        json={
            "employee_id": current_employee["id"],
            "issue_date": "2026-08-01",
            "condition_at_issue": "Good",
        },
    )
    assert allocated.status_code == 200, allocated.text

    manager_type = await _type_with(
        owner,
        ["Assets.View", "Assets.ManageStatus"],
        scope="company",
    )
    manager_user = await create_activated_user(owner, user_type_code=manager_type)
    operator_type = await _type_with(
        owner,
        ["Assets.View", permission],
        scope="company",
    )
    operator_user = await create_activated_user(owner, user_type_code=operator_type)

    async with await spawned_client() as manager:
        await authenticate(manager, manager_user["email"], "UserPass1!")
        controlled = await manager.post(
            f"/api/v1/assets/{asset['id']}/status",
            json={
                "status": controlled_status,
                "reason": f"H1 controlled state: {controlled_status}",
            },
        )
        assert controlled.status_code == 200, controlled.text

    before_detail = (await owner.get(f"/api/v1/assets/{asset['id']}")).json()
    before_history = (await owner.get(f"/api/v1/assets/{asset['id']}/history")).json()
    assert before_detail["status"] == controlled_status
    assert before_detail["currentAllocation"]["employeeId"] == current_employee["id"]

    if operation == "return":
        path = f"/api/v1/assets/{asset['id']}/return"
        payload = {
            "return_date": "2026-08-20",
            "return_condition": "Good",
            "remarks": "H1 denied return probe",
        }
    else:
        path = f"/api/v1/assets/{asset['id']}/transfer/employee"
        payload = {
            "employee_id": next_employee["id"],
            "transfer_date": "2026-08-20",
            "condition": "Good",
            "remarks": "H1 denied transfer probe",
        }

    async with await spawned_client() as operator:
        await authenticate(operator, operator_user["email"], "UserPass1!")
        denied = await operator.post(path, json=payload)
        assert denied.status_code == 409, denied.text
        assert denied.json()["error"]["code"] == "ASSET_STATUS_OPERATION_BLOCKED"
        assert denied.json()["error"]["details"] == [
            {
                "operation": "Return" if operation == "return" else "Employee Transfer",
                "currentStatus": controlled_status,
                "requiredStatus": "Allocated",
            }
        ]

    after_denial_detail = (await owner.get(f"/api/v1/assets/{asset['id']}")).json()
    after_denial_history = (
        await owner.get(f"/api/v1/assets/{asset['id']}/history")
    ).json()
    assert after_denial_detail == before_detail
    assert after_denial_history == before_history

    async with await spawned_client() as manager:
        await authenticate(manager, manager_user["email"], "UserPass1!")
        missing_reason = await manager.post(
            f"/api/v1/assets/{asset['id']}/status",
            json={"status": "Allocated", "reason": ""},
        )
        assert missing_reason.status_code == 422, missing_reason.text
        corrected = await manager.post(
            f"/api/v1/assets/{asset['id']}/status",
            json={
                "status": "Allocated",
                "reason": f"Authorized correction from {controlled_status}",
            },
        )
        assert corrected.status_code == 200, corrected.text
        assert corrected.json()["status"] == "Allocated"

    async with await spawned_client() as operator:
        await authenticate(operator, operator_user["email"], "UserPass1!")
        allowed = await operator.post(path, json=payload)
        assert allowed.status_code == 200, allowed.text
        if operation == "return":
            assert allowed.json()["status"] == "In Stock"
            assert allowed.json()["currentAllocation"] is None
        else:
            assert allowed.json()["status"] == "Allocated"
            assert allowed.json()["currentAllocation"]["employeeId"] == next_employee["id"]

    final_history = (await owner.get(f"/api/v1/assets/{asset['id']}/history")).json()
    correction = next(
        row
        for row in final_history["events"]
        if row["action"] == "asset.status.change"
        and row["newValues"] == {"status": "Allocated"}
    )
    assert correction["reason"] == f"Authorized correction from {controlled_status}"


@pytest.mark.asyncio
async def test_employee_profile_history_requires_audit_permission(
    client: AsyncClient,
) -> None:
    owner, _ = await owner_client(client)
    dxb = await office_id(owner, "DXB")
    employee = await _employee(owner, dxb)
    asset = await _create_pc(owner, dxb)
    assert (
        await owner.post(
            f"/api/v1/assets/{asset['id']}/allocate",
            json={
                "employee_id": employee["id"],
                "issue_date": "2026-08-01",
                "condition_at_issue": "Good",
            },
        )
    ).status_code == 200
    assert (
        await owner.post(
            f"/api/v1/assets/{asset['id']}/return",
            json={"return_date": "2026-08-20", "return_condition": "Good"},
        )
    ).status_code == 200

    view_type = await _type_with(owner, ["Assets.View"], scope="company")
    viewer = await create_activated_user(owner, user_type_code=view_type)
    async with await spawned_client() as restricted:
        await authenticate(restricted, viewer["email"], "UserPass1!")
        profile = await restricted.get(f"/api/v1/assets/employees/{employee['id']}")
        assert profile.status_code == 200, profile.text
        assert profile.json() == {"current": [], "history": []}
        direct_history = await restricted.get(f"/api/v1/assets/{asset['id']}/history")
        assert direct_history.status_code == 403
        assert direct_history.json()["error"]["details"] == [{"permission": "Assets.ViewAudit"}]


@pytest.mark.asyncio
async def test_office_scope_blocks_idor_tampering_audit_and_export_leakage(
    client: AsyncClient,
) -> None:
    owner, _ = await owner_client(client)
    dxb = await office_id(owner, "DXB")
    auh = await office_id(owner, "AUH")
    dxb_asset = await _create_pc(owner, dxb)
    hidden_serial = f"HIDDEN-{unique_tag()}".upper()
    auh_asset = await _create_pc(owner, auh, serial=hidden_serial)
    dxb_employee = await _employee(owner, dxb)
    auh_employee = await _employee(owner, auh)
    all_permissions = [
        "Assets.View",
        "Assets.ManageMaster",
        "Assets.ManageStock",
        "Assets.Allocate",
        "Assets.Transfer",
        "Assets.Return",
        "Assets.ManageStatus",
        "Assets.ViewAudit",
    ]
    office_type = await _type_with(owner, all_permissions, scope="office")
    office_admin = await create_activated_user(
        owner,
        user_type_code=office_type,
        office_id=dxb,
    )
    owner_history_before = await owner.get(f"/api/v1/assets/{auh_asset['id']}/history")
    before_event_ids = {row["id"] for row in owner_history_before.json()["events"]}

    async with await spawned_client() as restricted:
        await authenticate(restricted, office_admin["email"], "UserPass1!")
        listed = await restricted.get("/api/v1/assets")
        assert listed.status_code == 200
        listed_codes = {row["assetCode"] for row in listed.json()["items"]}
        assert dxb_asset["assetCode"] in listed_codes
        assert auh_asset["assetCode"] not in listed_codes
        assert hidden_serial not in listed.text

        hidden_probes = (
            ("GET", f"/api/v1/assets/{auh_asset['id']}", None),
            ("PATCH", f"/api/v1/assets/{auh_asset['id']}", {"brand": "Takeover"}),
            (
                "POST",
                f"/api/v1/assets/{auh_asset['id']}/identifiers",
                {"serial_number": "TAKEOVER", "reason": "Probe"},
            ),
            (
                "POST",
                f"/api/v1/assets/{auh_asset['id']}/condition",
                {"condition": "Fair", "reason": "Probe"},
            ),
            (
                "POST",
                f"/api/v1/assets/{auh_asset['id']}/allocate",
                {
                    "employee_id": dxb_employee["id"],
                    "issue_date": "2026-08-30",
                    "condition_at_issue": "Good",
                },
            ),
            (
                "POST",
                f"/api/v1/assets/{auh_asset['id']}/return",
                {"return_date": "2026-08-30", "return_condition": "Good"},
            ),
            (
                "POST",
                f"/api/v1/assets/{auh_asset['id']}/transfer/employee",
                {
                    "employee_id": dxb_employee["id"],
                    "transfer_date": "2026-08-30",
                    "condition": "Good",
                },
            ),
            (
                "POST",
                f"/api/v1/assets/{auh_asset['id']}/transfer/office",
                {"office_id": dxb, "transfer_date": "2026-08-30"},
            ),
            (
                "POST",
                f"/api/v1/assets/{auh_asset['id']}/status",
                {"status": "Lost", "reason": "Probe"},
            ),
            ("GET", f"/api/v1/assets/{auh_asset['id']}/history", None),
        )
        for method, path, body in hidden_probes:
            response = await restricted.request(method, path, json=body)
            assert response.status_code == 404, (method, path, response.text)
            assert response.json()["error"]["code"] == "ASSET_NOT_FOUND"

        hidden_profile = await restricted.get(f"/api/v1/assets/employees/{auh_employee['id']}")
        assert hidden_profile.status_code == 404
        assert hidden_profile.json()["error"]["code"] == "ASSET_EMPLOYEE_NOT_FOUND"
        hidden_audit = await restricted.get(
            "/api/v1/assets/audit", params={"assetId": auh_asset["id"]}
        )
        assert hidden_audit.status_code == 404
        audit_list = await restricted.get("/api/v1/assets/audit")
        assert auh_asset["id"] not in {row["entityId"] for row in audit_list.json()["items"]}

        report = await restricted.get(
            "/api/v1/assets/reports/asset_register", params={"officeId": auh}
        )
        assert report.status_code == 200
        assert report.json()["total"] == 0
        exported = await restricted.post(
            "/api/v1/assets/reports/export",
            json={"format": "xlsx", "report": "asset_register", "office_id": auh},
        )
        assert exported.status_code == 200
        assert hidden_serial.encode() not in exported.content

        pc = await _category(restricted, "PC")
        cross_office_create = await restricted.post(
            "/api/v1/assets",
            json=_pc_payload(pc["id"], auh),
        )
        assert cross_office_create.status_code == 404
        employee_tamper = await restricted.post(
            f"/api/v1/assets/{dxb_asset['id']}/allocate",
            json={
                "employee_id": auh_employee["id"],
                "issue_date": "2026-08-30",
                "condition_at_issue": "Good",
            },
        )
        assert employee_tamper.status_code == 404
        office_tamper = await restricted.post(
            f"/api/v1/assets/{dxb_asset['id']}/transfer/office",
            json={"office_id": auh, "transfer_date": "2026-08-30"},
        )
        assert office_tamper.status_code == 404
        category_takeover = await restricted.post(
            "/api/v1/assets/categories",
            json={"code": f"OFF-{unique_tag()[:6]}", "name": "Office global", "fields": []},
        )
        assert category_takeover.status_code == 403

    unchanged = await owner.get(f"/api/v1/assets/{auh_asset['id']}")
    assert unchanged.status_code == 200
    assert unchanged.json()["serialNumber"] == hidden_serial
    assert unchanged.json()["status"] == "In Stock"
    assert unchanged.json()["currentAllocation"] is None
    owner_history_after = await owner.get(f"/api/v1/assets/{auh_asset['id']}/history")
    assert {row["id"] for row in owner_history_after.json()["events"]} == before_event_ids


@pytest.mark.asyncio
async def test_mass_assignment_delete_audit_and_export_injection_controls(
    client: AsyncClient,
) -> None:
    owner, owner_user = await owner_client(client)
    dxb = await office_id(owner, "DXB")
    pc = await _category(owner, "PC")
    payload = _pc_payload(pc["id"], dxb)
    payload.update(
        {
            "brand": "=1+1",
            "model": "<script>alert(1)</script>",
            "asset_code": "AST-999999",
            "status": "Retired",
            "created_by_id": owner_user["id"],
        }
    )
    rejected = await owner.post("/api/v1/assets", json=payload)
    assert rejected.status_code == 422

    safe_payload = _pc_payload(pc["id"], dxb)
    safe_payload["brand"] = "=1+1"
    safe_payload["model"] = "<script>alert(1)</script>"
    created = await owner.post("/api/v1/assets", json=safe_payload)
    assert created.status_code == 200, created.text
    asset = created.json()
    mutation_attempts = (
        (
            "PATCH",
            f"/api/v1/assets/{asset['id']}",
            {"office_id": _DUMMY_ID, "status": "Lost", "serial_number": "OVERRIDE"},
        ),
        (
            "POST",
            f"/api/v1/assets/{asset['id']}/allocate",
            {
                "employee_id": owner_user["id"],
                "issue_date": "2026-08-30",
                "condition_at_issue": "Good",
                "issued_by_id": _DUMMY_ID,
            },
        ),
        (
            "POST",
            f"/api/v1/assets/{asset['id']}/status",
            {"status": "Lost", "reason": "Probe", "actor_id": _DUMMY_ID},
        ),
    )
    for method, path, body in mutation_attempts:
        response = await owner.request(method, path, json=body)
        assert response.status_code == 422, (path, response.text)

    changed = await owner.post(
        f"/api/v1/assets/{asset['id']}/status",
        json={"status": "Lost", "reason": "Security evidence"},
    )
    assert changed.status_code == 200
    history = await owner.get(f"/api/v1/assets/{asset['id']}/history")
    assert {row["action"] for row in history.json()["events"]} >= {
        "asset.create",
        "asset.status.change",
    }
    async with app.state.session_factory() as session:
        count = await session.scalar(
            select(func.count())
            .select_from(AuditEvent)
            .where(
                AuditEvent.entity_type == "asset",
                AuditEvent.entity_id == asset["id"],
                AuditEvent.actor_id == UUID(owner_user["id"]),
            )
        )
    assert count == 2

    xlsx = await owner.post(
        "/api/v1/assets/reports/export",
        json={"format": "xlsx", "report": "asset_register"},
    )
    assert xlsx.status_code == 200
    workbook = load_workbook(BytesIO(xlsx.content), data_only=False)
    rows = list(workbook["Asset Report"].iter_rows(values_only=True))
    headers = list(rows[0])
    brand_index = headers.index("Brand")
    matching = next(row for row in rows[1:] if row[0] == asset["assetCode"])
    assert matching[brand_index] == "'=1+1"
    printable = await owner.post(
        "/api/v1/assets/reports/export",
        json={"format": "print", "report": "asset_register"},
    )
    assert printable.status_code == 200
    assert "<script>alert(1)</script>" not in printable.text
    assert "&lt;script&gt;alert(1)&lt;/script&gt;" in printable.text
    csv = await owner.post(
        "/api/v1/assets/reports/export",
        json={"format": "csv", "report": "asset_register"},
    )
    assert csv.status_code == 422

    deleted = await owner.delete(f"/api/v1/assets/{asset['id']}")
    assert deleted.status_code == 405
    category_delete = await owner.delete(f"/api/v1/assets/categories/{pc['id']}")
    assert category_delete.status_code == 405
    still_exists = await owner.get(f"/api/v1/assets/{asset['id']}")
    assert still_exists.status_code == 200

from __future__ import annotations

from datetime import UTC, datetime, time, timedelta
from io import BytesIO
from uuid import UUID

import pytest
from helpers import (
    authenticate,
    business_today,
    create_activated_user,
    create_product_variant,
    office_id,
    owner_client,
    spawned_client,
    utc_today,
    unique_tag,
)
from httpx import AsyncClient
from nexa_bos_api.identity.models import AuditEvent
from nexa_bos_api.main import app
from openpyxl import load_workbook
from sqlalchemy import select, text
from test_applications import _catalog, _create_app, _customer, _ensure_test_workflow
from test_role_readiness import _configure_system_type
from test_tat_delay import _move


async def _reporting_user(
    authed: AsyncClient,
    *,
    scope: str | None,
    permissions: list[str],
    office_id: str | None = None,
    manager_id: str | None = None,
    password: str = "UserPass1!",
    can_be_case_owner: bool = True,
    can_be_reporting_manager: bool = False,
) -> dict:
    tag = unique_tag().upper()
    created = await authed.post(
        "/api/v1/user-types",
        json={
            "name": f"RPT {tag}",
            "code": f"R{tag[:8]}",
            "can_be_case_owner": can_be_case_owner,
            "can_be_reporting_manager": can_be_reporting_manager,
        },
    )
    assert created.status_code == 200, created.text
    type_id = created.json()["id"]
    await authed.post(f"/api/v1/user-types/{type_id}/activate")
    await authed.put(
        f"/api/v1/user-types/{type_id}/permissions",
        json={"permissions": permissions},
    )
    await authed.put(
        f"/api/v1/user-types/{type_id}/reporting-scope",
        json={"reporting_visibility_scope": scope},
    )
    return await create_activated_user(
        authed,
        user_type_code=created.json()["code"],
        password=password,
        office_id=office_id,
        manager_id=manager_id,
    )


async def _submit_and_fund(client: AsyncClient, application: dict, amount: str = "8000") -> dict:
    submitted = await client.post(
        f"/api/v1/applications/{application['id']}/case-number",
        json={"bank_case_number": f"RPT-{unique_tag()[:8]}"},
    )
    assert submitted.status_code == 200, submitted.text
    await _move(client, application, "approved", {"approved_amount": amount})
    await _move(client, application, "booked", {"booked_amount": amount})
    return await _move(client, application, "fund_released", {"funded_amount": amount})


VIEW_PERMS = ["Dashboard.View", "Reports.View"]
EXPORT_PERMS = [
    "Dashboard.View",
    "Reports.View",
    "Reports.ExportExcel",
    "Reports.ExportPDF",
    "Reports.Print",
]


@pytest.mark.asyncio
async def test_dashboard_requires_permission(client: AsyncClient) -> None:
    authed, _owner = await owner_client(client)
    user = await _reporting_user(authed, scope="company", permissions=["Users.View"])
    async with await spawned_client() as other:
        await authenticate(other, user["email"], "UserPass1!")
        response = await other.get("/api/v1/reports/dashboard")
        assert response.status_code == 403


@pytest.mark.asyncio
async def test_permission_without_reporting_scope_returns_no_data(client: AsyncClient) -> None:
    authed, owner = await owner_client(client)
    dib, _eib, pf, _cc = await _catalog(authed)
    customer = await _customer(authed, "NoScope")
    created = await _create_app(
        authed,
        customer_id=customer["id"],
        bank_id=dib["id"],
        product_id=pf["id"],
        case_owner_id=owner["id"],
        requested_amount="5000",
    )
    await _submit_and_fund(authed, created)
    user = await _reporting_user(authed, scope=None, permissions=VIEW_PERMS)
    async with await spawned_client() as other:
        await authenticate(other, user["email"], "UserPass1!")
        dashboard = await other.get("/api/v1/reports/dashboard")
        assert dashboard.status_code == 200, dashboard.text
        body = dashboard.json()
        assert body["reportingScope"] is None
        assert body["kpis"]["submitted"]["count"] == 0
        assert body["kpis"]["funded"]["count"] == 0
        assert body["rankings"]["employees"] == []
        drill = await other.get("/api/v1/reports/applications?metric=funded")
        assert drill.json()["total"] == 0
        assert drill.json()["items"] == []


@pytest.mark.asyncio
async def test_owner_company_wide_dashboard_and_mtd_default(client: AsyncClient) -> None:
    authed, owner = await owner_client(client)
    dib, _eib, pf, _cc = await _catalog(authed)
    customer = await _customer(authed, "OwnerDash")
    created = await _create_app(
        authed,
        customer_id=customer["id"],
        bank_id=dib["id"],
        product_id=pf["id"],
        case_owner_id=owner["id"],
        requested_amount="12000",
    )
    await _submit_and_fund(authed, created, "11000")
    dashboard = await authed.get("/api/v1/reports/dashboard")
    assert dashboard.status_code == 200, dashboard.text
    body = dashboard.json()
    assert body["period"]["key"] == "mtd"
    assert body["reportingScope"] == "Company-wide"
    assert body["currency"] == "AED"
    assert body["kpis"]["submitted"]["count"] >= 1
    assert body["kpis"]["funded"]["count"] >= 1
    assert body["kpis"]["personalFinance"]["count"] >= 1
    assert body["conversions"]["submittedToApproved"] is not None
    drill = await authed.get("/api/v1/reports/applications?metric=funded")
    drill_items = drill.json()["items"]
    codes = {item["applicationCode"] for item in drill_items}
    assert created["applicationCode"] in codes
    drill_item = next(
        item for item in drill_items if item["applicationCode"] == created["applicationCode"]
    )
    assert drill_item["bankName"] == created["bankName"]
    assert drill_item["productName"] == created["productName"]
    assert drill_item["productVariantId"] == created["productVariantId"]
    assert drill_item["productVariantCode"] == created["productVariantCode"]


@pytest.mark.asyncio
async def test_own_office_team_scopes_and_query_param_cannot_bypass(client: AsyncClient) -> None:
    authed, _owner = await owner_client(client)
    dib, _eib, pf, _cc = await _catalog(authed)
    dxb = await office_id(authed, "DXB")
    auh = await office_id(authed, "AUH")
    manager = await _reporting_user(
        authed,
        scope="team",
        permissions=VIEW_PERMS,
        office_id=dxb,
        can_be_reporting_manager=True,
    )
    own_user = await _reporting_user(authed, scope="own", permissions=VIEW_PERMS, office_id=dxb)
    office_user = await _reporting_user(
        authed, scope="office", permissions=VIEW_PERMS, office_id=dxb
    )
    report = await _reporting_user(
        authed, scope="team", permissions=VIEW_PERMS, office_id=dxb, manager_id=manager["id"]
    )
    other_office = await _reporting_user(authed, scope="own", permissions=VIEW_PERMS, office_id=auh)
    own_app = await _create_app(
        authed,
        customer_id=(await _customer(authed, "OwnApp"))["id"],
        bank_id=dib["id"],
        product_id=pf["id"],
        case_owner_id=own_user["id"],
        requested_amount="3000",
    )
    await _submit_and_fund(authed, own_app, "3000")
    team_app = await _create_app(
        authed,
        customer_id=(await _customer(authed, "TeamApp"))["id"],
        bank_id=dib["id"],
        product_id=pf["id"],
        case_owner_id=report["id"],
        requested_amount="4000",
    )
    await _submit_and_fund(authed, team_app, "4000")
    auh_app = await _create_app(
        authed,
        customer_id=(await _customer(authed, "AuhApp"))["id"],
        bank_id=dib["id"],
        product_id=pf["id"],
        case_owner_id=other_office["id"],
        requested_amount="9000",
    )
    await _submit_and_fund(authed, auh_app, "9000")

    async with await spawned_client() as scoped:
        await authenticate(scoped, own_user["email"], "UserPass1!")
        own_dash = (await scoped.get("/api/v1/reports/dashboard")).json()
        assert own_dash["reportingScope"] == "Own Performance"
        own_codes = {
            item["applicationCode"]
            for item in (await scoped.get("/api/v1/reports/applications?metric=funded")).json()[
                "items"
            ]
        }
        assert own_app["applicationCode"] in own_codes
        assert team_app["applicationCode"] not in own_codes
        assert auh_app["applicationCode"] not in own_codes
        bypass = await scoped.get(
            f"/api/v1/reports/dashboard?employee_id={other_office['id']}&reporting_scope=company"
        )
        assert bypass.json()["kpis"]["funded"]["count"] == 0
        profile = await scoped.get(f"/api/v1/reports/employees/{other_office['id']}")
        assert profile.status_code == 404

    async with await spawned_client() as scoped:
        await authenticate(scoped, office_user["email"], "UserPass1!")
        codes = {
            item["applicationCode"]
            for item in (await scoped.get("/api/v1/reports/applications?metric=funded")).json()[
                "items"
            ]
        }
        assert own_app["applicationCode"] in codes
        assert team_app["applicationCode"] in codes
        assert auh_app["applicationCode"] not in codes
        options = (await scoped.get("/api/v1/reports/filters")).json()
        employee_ids = {item["id"] for item in options["employees"]}
        assert other_office["id"] not in employee_ids

    async with await spawned_client() as scoped:
        await authenticate(scoped, manager["email"], "UserPass1!")
        codes = {
            item["applicationCode"]
            for item in (await scoped.get("/api/v1/reports/applications?metric=funded")).json()[
                "items"
            ]
        }
        assert team_app["applicationCode"] in codes
        assert auh_app["applicationCode"] not in codes


@pytest.mark.asyncio
async def test_se_dashboard_is_personal_and_includes_read_only_attendance_and_targets(
    client: AsyncClient,
) -> None:
    authed, _owner = await owner_client(client)
    dxb = await office_id(authed, "DXB")
    dib, _eib, pf, _cc = await _catalog(authed)
    await _ensure_test_workflow(authed, dib["id"], pf["id"])
    variant = await create_product_variant(authed, bank_id=dib["id"], product_id=pf["id"])
    await _configure_system_type(
        authed,
        "SE",
        permissions=["Dashboard.View", "Applications.View", "Applications.Create"],
        application_scope="own",
        reporting_scope="company",
        can_be_case_owner=True,
    )
    first = await create_activated_user(authed, user_type_code="SE", office_id=dxb)
    second = await create_activated_user(authed, user_type_code="SE", office_id=dxb)

    async def create_owned(user: dict, label: str) -> dict:
        async with await spawned_client() as scoped:
            await authenticate(scoped, user["email"], "UserPass1!")
            created = await scoped.post(
                "/api/v1/applications",
                json={
                    "customer": {
                        "customer_type": "individual",
                        "full_name": f"Dashboard {label} {unique_tag()}",
                        "mobile": f"+9715{unique_tag()[:8]}",
                    },
                    "bank_id": dib["id"],
                    "product_id": pf["id"],
                    "product_variant_id": variant["id"],
                    "requested_amount": "9000",
                },
            )
            assert created.status_code == 200, created.text
            return created.json()

    own_application = await create_owned(first, "Own")
    other_application = await create_owned(second, "Other")
    target = await authed.post(
        "/api/v1/targets",
        json={
            "level": "employee",
            "entity_id": first["id"],
            "period_month": business_today().replace(day=1).isoformat(),
            "product_id": pf["id"],
            "milestone": "submitted",
            "measurement": "count",
            "target_value": "4",
        },
    )
    assert target.status_code == 200, target.text
    activated = await authed.post(f"/api/v1/targets/{target.json()['id']}/activate")
    assert activated.status_code == 200, activated.text
    attendance = await authed.put(
        "/api/v1/attendance/records",
        json={
            "attendance_date": business_today().isoformat(),
            "entries": [
                {
                    "employee_id": first["id"],
                    "status": "Present",
                    "time_in": "09:05",
                    "time_out": "17:05",
                }
            ],
        },
    )
    assert attendance.status_code == 200, attendance.text

    async with await spawned_client() as scoped:
        await authenticate(scoped, first["email"], "UserPass1!")
        dashboard = await scoped.get(
            f"/api/v1/reports/dashboard?period=mtd&employee_id={second['id']}&office_id={dxb}"
        )
        assert dashboard.status_code == 200, dashboard.text
        body = dashboard.json()
        workspace = body["seWorkspace"]
        assert workspace["kpis"]["applications"]["count"] == 1
        recent_codes = {item["localFileNumber"] for item in workspace["recentApplications"]}
        assert own_application["applicationCode"] in recent_codes
        assert other_application["applicationCode"] not in recent_codes
        assert workspace["targetProgress"]["assigned"] == "4.00"
        assert body["personalAttendance"]["today"]["status"] == "Present"
        assert body["personalAttendance"]["today"]["workedMinutes"] == 480
        assert body["personalPerformance"]["applicationMetrics"] is not None
        assert len(workspace["trend"]) == 6
        assert workspace["stages"]
        assert workspace["products"][0]["code"] == pf["code"]

        filtered = await scoped.get(
            "/api/v1/applications?dashboard_metric=applications&dashboard_period=mtd&page_size=50"
        )
        assert filtered.status_code == 200, filtered.text
        filtered_codes = {item["applicationCode"] for item in filtered.json()["items"]}
        assert own_application["applicationCode"] in filtered_codes
        assert other_application["applicationCode"] not in filtered_codes


@pytest.mark.asyncio
async def test_cod_dashboard_is_office_scoped_operational_and_personal(
    client: AsyncClient,
) -> None:
    authed, _owner = await owner_client(client)
    dxb = await office_id(authed, "DXB")
    auh = await office_id(authed, "AUH")
    dib, _eib, pf, _cc = await _catalog(authed)
    await _ensure_test_workflow(authed, dib["id"], pf["id"])
    variant = await create_product_variant(authed, bank_id=dib["id"], product_id=pf["id"])
    await _configure_system_type(
        authed,
        "TL",
        permissions=["Applications.View", "Applications.Edit"],
        application_scope="team",
    )

    async def review_team(office):
        tag = unique_tag()
        department = await authed.post(
            "/api/v1/departments",
            json={
                "office_id": office,
                "code": f"CD{tag}",
                "name": f"COD review {tag}",
            },
        )
        assert department.status_code == 200, department.text
        team = await authed.post(
            "/api/v1/teams",
            json={
                "office_id": office,
                "department_id": department.json()["id"],
                "code": f"CT{tag}",
                "name": f"COD team {tag}",
            },
        )
        assert team.status_code == 200, team.text
        return {"department_id": department.json()["id"], "team_id": team.json()["id"]}

    dxb_review_team = await review_team(dxb)
    auh_review_team = await review_team(auh)
    await _configure_system_type(
        authed,
        "COD",
        permissions=[
            "Dashboard.View",
            "Applications.View",
            "Applications.Submit",
            "Applications.UpdateStage",
            "Applications.MarkDelay",
        ],
        application_scope="office",
        reporting_scope="office",
        can_be_case_owner=True,
    )
    await _configure_system_type(
        authed,
        "SE",
        permissions=["Dashboard.View", "Applications.View", "Applications.Create"],
        application_scope="own",
        reporting_scope="own",
        can_be_case_owner=True,
    )
    dxb_sm = await create_activated_user(authed, user_type_code="SM", office_id=dxb)
    dxb_cod = await create_activated_user(
        authed, user_type_code="COD", office_id=dxb, manager_id=dxb_sm["id"]
    )
    dxb_tl = await create_activated_user(
        authed, user_type_code="TL", office_id=dxb, manager_id=dxb_cod["id"], **dxb_review_team
    )
    dxb_se = await create_activated_user(
        authed, user_type_code="SE", office_id=dxb, manager_id=dxb_tl["id"], **dxb_review_team
    )
    auh_sm = await create_activated_user(authed, user_type_code="SM", office_id=auh)
    auh_cod = await create_activated_user(
        authed, user_type_code="COD", office_id=auh, manager_id=auh_sm["id"]
    )
    auh_tl = await create_activated_user(
        authed, user_type_code="TL", office_id=auh, manager_id=auh_cod["id"], **auh_review_team
    )
    auh_se = await create_activated_user(
        authed, user_type_code="SE", office_id=auh, manager_id=auh_tl["id"], **auh_review_team
    )

    async def create_owned(user: dict, label: str) -> dict:
        async with await spawned_client() as scoped:
            await authenticate(scoped, user["email"], "UserPass1!")
            response = await scoped.post(
                "/api/v1/applications",
                json={
                    "customer": {
                        "customer_type": "individual",
                        "full_name": f"COD Dashboard {label} {unique_tag()}",
                        "mobile": f"+9715{unique_tag()[:8]}",
                    },
                    "bank_id": dib["id"],
                    "product_id": pf["id"],
                    "product_variant_id": variant["id"],
                    "requested_amount": "10000",
                },
            )
            assert response.status_code == 200, response.text
            return response.json()

    dxb_waiting = await create_owned(dxb_se, "DXB waiting")
    dxb_processing = await create_owned(dxb_se, "DXB processing")
    auh_waiting = await create_owned(auh_se, "AUH waiting")
    for leader, applications in ((dxb_tl, (dxb_waiting, dxb_processing)), (auh_tl, (auh_waiting,))):
        async with await spawned_client() as reviewer:
            await authenticate(reviewer, leader["email"], "UserPass1!")
            for application in applications:
                path = f"/api/v1/applications/{application['id']}/internal-review"
                state = (await reviewer.get(path)).json()
                forwarded = await reviewer.post(
                    path, json={"action": "forward", "expected_event_id": state["eventId"]}
                )
                assert forwarded.status_code == 200, forwarded.text

    async with await spawned_client() as dxb_client:
        await authenticate(dxb_client, dxb_cod["email"], "UserPass1!")
        submitted = await dxb_client.post(
            f"/api/v1/applications/{dxb_processing['id']}/case-number",
            json={"bank_case_number": f"COD-{unique_tag()[:8]}"},
        )
        assert submitted.status_code == 200, submitted.text
        delay = await dxb_client.post(
            f"/api/v1/applications/{dxb_processing['id']}/delays",
            json={"delay_type": "Customer", "reason": "Disposable dashboard delay"},
        )
        assert delay.status_code == 200, delay.text

    target = await authed.post(
        "/api/v1/targets",
        json={
            "level": "employee",
            "entity_id": dxb_cod["id"],
            "period_month": business_today().replace(day=1).isoformat(),
            "product_id": pf["id"],
            "milestone": "submitted",
            "measurement": "count",
            "target_value": "3",
        },
    )
    assert target.status_code == 200, target.text
    assert (await authed.post(f"/api/v1/targets/{target.json()['id']}/activate")).status_code == 200
    attendance = await authed.put(
        "/api/v1/attendance/records",
        json={
            "attendance_date": business_today().isoformat(),
            "entries": [
                {
                    "employee_id": dxb_cod["id"],
                    "status": "Present",
                    "time_in": "09:00",
                    "time_out": "17:00",
                }
            ],
        },
    )
    assert attendance.status_code == 200, attendance.text

    async with await spawned_client() as dxb_client:
        await authenticate(dxb_client, dxb_cod["email"], "UserPass1!")
        response = await dxb_client.get(f"/api/v1/reports/dashboard?period=mtd&office_id={auh}")
        assert response.status_code == 200, response.text
        body = response.json()
        workspace = body["codWorkspace"]
        assert workspace["office"]["id"] == dxb
        assert workspace["office"]["scope"] == "Office operations"
        assert workspace["kpis"]["newCases"] >= 2
        assert workspace["kpis"]["awaitingSubmission"] >= 1
        assert workspace["kpis"]["submitted"] >= 1
        assert workspace["kpis"]["delayed"] >= 1
        recent = {item["localFileNumber"] for item in workspace["queues"]["recentUpdates"]}
        assert dxb_waiting["applicationCode"] in recent
        assert dxb_processing["applicationCode"] in recent
        assert auh_waiting["applicationCode"] not in recent
        staff = {row["id"]: row for row in workspace["staff"]}
        assert staff[dxb_tl["id"]]["downline"] is True
        assert staff[dxb_se["id"]]["downline"] is True
        assert staff[dxb_sm["id"]]["downline"] is False
        assert workspace["activity"] == {
            "reviewed": 1,
            "submitted": 1,
            "stageUpdates": 0,
        }
        assert body["personalPerformance"]["target"]["assigned"] == "3.00"
        assert body["personalAttendance"]["today"]["workedMinutes"] == 480
        assert body["seWorkspace"] is None
        delayed_rows = await dxb_client.get(
            "/api/v1/applications?dashboard_metric=delayed&dashboard_period=mtd&page_size=50"
        )
        assert delayed_rows.status_code == 200, delayed_rows.text
        delayed_codes = {item["applicationCode"] for item in delayed_rows.json()["items"]}
        assert dxb_processing["applicationCode"] in delayed_codes
        assert auh_waiting["applicationCode"] not in delayed_codes
        cross_office = await dxb_client.get(f"/api/v1/applications/{auh_waiting['id']}")
        assert cross_office.status_code == 404

    async with await spawned_client() as auh_client:
        await authenticate(auh_client, auh_cod["email"], "UserPass1!")
        workspace = (await auh_client.get("/api/v1/reports/dashboard?period=mtd")).json()[
            "codWorkspace"
        ]
        recent = {item["localFileNumber"] for item in workspace["queues"]["recentUpdates"]}
        assert auh_waiting["applicationCode"] in recent
        assert dxb_waiting["applicationCode"] not in recent
        assert dxb_processing["applicationCode"] not in recent


@pytest.mark.asyncio
async def test_non_application_role_personal_dashboard_never_fakes_sales_metrics(
    client: AsyncClient,
) -> None:
    authed, _owner = await owner_client(client)
    user = await _reporting_user(
        authed,
        scope=None,
        permissions=["Dashboard.View"],
        can_be_case_owner=False,
    )
    async with await spawned_client() as scoped:
        await authenticate(scoped, user["email"], "UserPass1!")
        response = await scoped.get("/api/v1/reports/dashboard")
        assert response.status_code == 200, response.text
        body = response.json()
        assert body["personalPerformance"]["applicationMetrics"] is None
        assert body["personalPerformance"]["target"]["count"] == 0
        assert body["personalAttendance"]["today"]["date"] == business_today().isoformat()
        assert body["seWorkspace"] is None


@pytest.mark.asyncio
async def test_event_time_attribution_survives_reassignment(client: AsyncClient) -> None:
    authed, _owner = await owner_client(client)
    dib, _eib, pf, _cc = await _catalog(authed)
    first = await _reporting_user(authed, scope="own", permissions=VIEW_PERMS)
    second = await _reporting_user(authed, scope="own", permissions=VIEW_PERMS)
    created = await _create_app(
        authed,
        customer_id=(await _customer(authed, "ReassignPerf"))["id"],
        bank_id=dib["id"],
        product_id=pf["id"],
        case_owner_id=first["id"],
        requested_amount="7000",
    )
    await authed.post(
        f"/api/v1/applications/{created['id']}/case-number",
        json={"bank_case_number": f"RAS-{unique_tag()[:8]}"},
    )
    await _move(authed, created, "approved", {"approved_amount": "7000"})
    reassigned = await authed.post(
        f"/api/v1/applications/{created['id']}/reassign-owner",
        json={"case_owner_id": second["id"], "reason": "Coverage"},
    )
    assert reassigned.status_code == 200, reassigned.text
    await _move(authed, created, "booked", {"booked_amount": "7000"})
    await _move(authed, created, "fund_released", {"funded_amount": "7000"})

    async with await spawned_client() as scoped:
        await authenticate(scoped, first["email"], "UserPass1!")
        first_dash = (await scoped.get("/api/v1/reports/dashboard")).json()
        assert first_dash["kpis"]["submitted"]["count"] >= 1
        assert first_dash["kpis"]["approved"]["count"] >= 1
        assert first_dash["kpis"]["booked"]["count"] == 0
        assert first_dash["kpis"]["funded"]["count"] == 0
        submitted = (await scoped.get("/api/v1/reports/applications?metric=submitted")).json()
        assert created["applicationCode"] in {
            item["applicationCode"] for item in submitted["items"]
        }
        funded = (await scoped.get("/api/v1/reports/applications?metric=funded")).json()
        assert created["applicationCode"] not in {
            item["applicationCode"] for item in funded["items"]
        }

    async with await spawned_client() as scoped:
        await authenticate(scoped, second["email"], "UserPass1!")
        second_dash = (await scoped.get("/api/v1/reports/dashboard")).json()
        assert second_dash["kpis"]["submitted"]["count"] == 0
        assert second_dash["kpis"]["booked"]["count"] >= 1
        assert second_dash["kpis"]["funded"]["count"] >= 1


@pytest.mark.asyncio
async def test_pending_is_current_state_and_can_include_earlier_created(
    client: AsyncClient,
) -> None:
    authed, owner = await owner_client(client)
    dib, _eib, pf, _cc = await _catalog(authed)
    pending_app = await _create_app(
        authed,
        customer_id=(await _customer(authed, "PendingOld"))["id"],
        bank_id=dib["id"],
        product_id=pf["id"],
        case_owner_id=owner["id"],
        requested_amount="2000",
    )
    earlier = utc_today().replace(day=1) - timedelta(days=10)
    async with app.state.session_factory() as session:
        await session.execute(
            text("UPDATE applications SET created_at = :created_at WHERE id = :id"),
            {
                "created_at": datetime.combine(earlier, time.min, tzinfo=UTC),
                "id": UUID(pending_app["id"]),
            },
        )
        await session.commit()
    dashboard = (await authed.get("/api/v1/reports/dashboard?period=mtd")).json()
    assert dashboard["kpis"]["pending"]["count"] >= 1
    pending_codes: set[str] = set()
    first_page = (
        await authed.get("/api/v1/reports/applications?metric=pending&period=mtd&page_size=50")
    ).json()
    pending_codes.update(item["applicationCode"] for item in first_page["items"])
    for page in range(2, first_page["pagination"]["totalPages"] + 1):
        response = await authed.get(
            f"/api/v1/reports/applications?metric=pending&period=mtd&page_size=50&page={page}"
        )
        assert response.status_code == 200, response.text
        pending_codes.update(item["applicationCode"] for item in response.json()["items"])
    assert pending_app["applicationCode"] in pending_codes
    stages = {row["name"] for row in dashboard["stageBreakdown"]}
    assert stages


@pytest.mark.asyncio
async def test_conversions_rankings_ties_comparisons_custom_and_since_joining(
    client: AsyncClient,
) -> None:
    authed, owner = await owner_client(client)
    dib, eib, pf, cc = await _catalog(authed)
    first = await _reporting_user(authed, scope="company", permissions=VIEW_PERMS)
    second = await _reporting_user(authed, scope="company", permissions=VIEW_PERMS)
    for owner_id, amount in ((first["id"], "5000"), (second["id"], "5000")):
        app = await _create_app(
            authed,
            customer_id=(await _customer(authed, "Tie"))["id"],
            bank_id=dib["id"],
            product_id=pf["id"],
            case_owner_id=owner_id,
            requested_amount=amount,
        )
        await _submit_and_fund(authed, app, amount)
    cc_app = await _create_app(
        authed,
        customer_id=(await _customer(authed, "CCCount"))["id"],
        bank_id=dib["id"],
        product_id=cc["id"],
        case_owner_id=first["id"],
        requested_amount=None,
    )
    await authed.post(
        f"/api/v1/applications/{cc_app['id']}/case-number",
        json={"bank_case_number": f"CC-{unique_tag()[:8]}"},
    )
    rejected = await _create_app(
        authed,
        customer_id=(await _customer(authed, "Reject"))["id"],
        bank_id=eib["id"],
        product_id=pf["id"],
        case_owner_id=first["id"],
        requested_amount="1000",
    )
    await authed.post(
        f"/api/v1/applications/{rejected['id']}/outcome",
        json={"outcome": "Final Rejected", "reason": "Policy"},
    )
    empty = (await authed.get("/api/v1/reports/dashboard?period=previous_month")).json()
    assert empty["conversions"]["submittedToApproved"] is None
    rankings = (await authed.get("/api/v1/reports/rankings?ranking_metric=funded_value")).json()[
        "rankings"
    ]
    employee_rows = [
        row for row in rankings["employees"] if row["id"] in {first["id"], second["id"]}
    ]
    assert len(employee_rows) == 2
    assert employee_rows[0]["rank"] == employee_rows[1]["rank"]
    dashboard = (await authed.get("/api/v1/reports/dashboard")).json()
    assert dashboard["kpis"]["creditCard"]["count"] >= 1
    assert dashboard["kpis"]["creditCard"]["value"] is None
    assert dashboard["kpis"]["finalRejected"]["count"] >= 1
    today = utc_today().isoformat()
    custom = await authed.get(
        f"/api/v1/reports/dashboard?period=custom&date_from={today}&date_to={today}"
    )
    assert custom.status_code == 200, custom.text
    assert custom.json()["period"]["key"] == "custom"
    profile = await authed.get(f"/api/v1/reports/employees/{first['id']}?period=since_joining")
    assert profile.status_code == 200, profile.text
    assert profile.json()["period"]["key"] == "since_joining"
    assert profile.json()["employee"]["employeeCode"] == first["employeeCode"]
    compare = await authed.get(
        "/api/v1/reports/comparisons?kind=period&period=month&metric=funded_value"
    )
    assert compare.status_code == 200, compare.text
    assert compare.json()["percentageChange"] is None or isinstance(
        compare.json()["percentageChange"], float
    )
    entity = await authed.get(
        "/api/v1/reports/comparisons"
        f"?kind=entity&dimension=employee&left_id={first['id']}&right_id={second['id']}"
        "&metric=funded_value&period=mtd"
    )
    assert entity.status_code == 200, entity.text
    assert entity.json()["absoluteDifference"] is not None
    missing = await authed.get("/api/v1/reports/dashboard?period=custom")
    assert missing.status_code == 422


@pytest.mark.asyncio
async def test_active_delay_drilldown_and_exports(client: AsyncClient) -> None:
    authed, owner = await owner_client(client)
    dib, _eib, pf, _cc = await _catalog(authed)
    delayed = await _create_app(
        authed,
        customer_id=(await _customer(authed, "DelayDash"))["id"],
        bank_id=dib["id"],
        product_id=pf["id"],
        case_owner_id=owner["id"],
        requested_amount="1500",
    )
    marked = await authed.post(
        f"/api/v1/applications/{delayed['id']}/delays",
        json={"delay_type": "Bank", "reason": "Waiting on bank"},
    )
    assert marked.status_code == 200, marked.text
    dashboard = (await authed.get("/api/v1/reports/dashboard")).json()
    assert dashboard["activeDelays"]["Bank"] >= 1
    delay_rows = (await authed.get("/api/v1/reports/applications?metric=delay_bank")).json()
    assert delayed["applicationCode"] in {item["applicationCode"] for item in delay_rows["items"]}
    excel = await authed.post(
        "/api/v1/reports/export",
        json={"format": "xlsx", "report": "dashboard", "period": "mtd"},
    )
    assert excel.status_code == 200, excel.text
    assert "spreadsheetml" in excel.headers["content-type"]
    pdf = await authed.post(
        "/api/v1/reports/export",
        json={"format": "pdf", "report": "drill_down", "period": "mtd", "metric": "pending"},
    )
    assert pdf.status_code == 200, pdf.text
    assert pdf.headers["content-type"] == "application/pdf"
    printed = await authed.post(
        "/api/v1/reports/export",
        json={"format": "print", "report": "rankings", "period": "mtd"},
    )
    assert printed.status_code == 200, printed.text
    assert "text/html" in printed.headers["content-type"]
    async with app.state.session_factory() as session:
        events = (
            (await session.execute(select(AuditEvent).where(AuditEvent.action == "reports.export")))
            .scalars()
            .all()
        )
        assert events
        assert events[-1].new_values["exportType"] in {"xlsx", "pdf", "print"}
        assert "file" not in (events[-1].new_values or {})
    limited = await _reporting_user(authed, scope="company", permissions=VIEW_PERMS)
    async with await spawned_client() as other:
        await authenticate(other, limited["email"], "UserPass1!")
        denied_excel = await other.post(
            "/api/v1/reports/export",
            json={"format": "xlsx", "report": "dashboard", "period": "mtd"},
        )
        assert denied_excel.status_code == 403
        denied_pdf = await other.post(
            "/api/v1/reports/export",
            json={"format": "pdf", "report": "dashboard", "period": "mtd"},
        )
        assert denied_pdf.status_code == 403
        denied_print = await other.post(
            "/api/v1/reports/export",
            json={"format": "print", "report": "dashboard", "period": "mtd"},
        )
        assert denied_print.status_code == 403
    exporter = await _reporting_user(authed, scope="own", permissions=EXPORT_PERMS)
    own_app = await _create_app(
        authed,
        customer_id=(await _customer(authed, "ExportOwn"))["id"],
        bank_id=dib["id"],
        product_id=pf["id"],
        case_owner_id=exporter["id"],
        requested_amount="2500",
    )
    variant_injection = '=HYPERLINK("https://invalid.example","<script>alert(1)</script>")'
    renamed_variant = await authed.patch(
        f"/api/v1/product-variants/{own_app['productVariantId']}",
        json={"name": variant_injection, "description": "Export safety coverage"},
    )
    assert renamed_variant.status_code == 200, renamed_variant.text
    await _submit_and_fund(authed, own_app, "2500")
    async with await spawned_client() as other:
        await authenticate(other, exporter["email"], "UserPass1!")
        exported = await other.post(
            "/api/v1/reports/export",
            json={"format": "xlsx", "report": "drill_down", "period": "mtd", "metric": "funded"},
        )
        assert exported.status_code == 200, exported.text
        workbook = load_workbook(BytesIO(exported.content), read_only=True, data_only=False)
        result_rows = list(workbook["Results"].iter_rows(values_only=True))
        headers = list(result_rows[0])
        assert "productVariantCode" in headers
        assert "productVariantName" in headers
        application_index = headers.index("applicationCode")
        variant_index = headers.index("productVariantCode")
        variant_name_index = headers.index("productVariantName")
        own_export_row = next(
            row for row in result_rows[1:] if row[application_index] == own_app["applicationCode"]
        )
        assert own_export_row[variant_index] == own_app["productVariantCode"]
        assert own_export_row[variant_name_index] == f"'{variant_injection}"
        printed_drill = await other.post(
            "/api/v1/reports/export",
            json={"format": "print", "report": "drill_down", "period": "mtd", "metric": "funded"},
        )
        assert printed_drill.status_code == 200, printed_drill.text
        assert "<script>alert(1)</script>" not in printed_drill.text
        assert "&lt;script&gt;alert(1)&lt;/script&gt;" in printed_drill.text
        scoped_drill = (await other.get("/api/v1/reports/applications?metric=funded")).json()
        assert all(
            item["applicationCode"] != delayed["applicationCode"] for item in scoped_drill["items"]
        )
        assert own_app["applicationCode"] in {
            item["applicationCode"] for item in scoped_drill["items"]
        }

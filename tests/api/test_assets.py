from __future__ import annotations

import asyncio
from datetime import date
from io import BytesIO

import pytest
from helpers import (
    authenticate,
    create_activated_user,
    designation_id,
    office_id,
    owner_client,
    spawned_client,
    unique_tag,
)
from httpx import AsyncClient
from openpyxl import load_workbook


def _digits(length: int) -> str:
    return str(int(unique_tag(), 16)).zfill(length)[-length:]


async def _category(client: AsyncClient, code: str) -> dict:
    response = await client.get("/api/v1/assets/categories")
    assert response.status_code == 200, response.text
    return next(item for item in response.json()["items"] if item["code"] == code)


def _pc_payload(category_id: str, office: str, *, serial: str | None = None) -> dict:
    return {
        "category_id": category_id,
        "office_id": office,
        "condition": "New",
        "brand": "Dell",
        "model": "Latitude 7450",
        "serial_number": serial or f"PC-{unique_tag()}",
        "attributes": {},
        "description": "Task 13 tracked computer",
    }


async def _create_pc(
    client: AsyncClient,
    office: str,
    *,
    serial: str | None = None,
) -> dict:
    pc = await _category(client, "PC")
    response = await client.post(
        "/api/v1/assets",
        json=_pc_payload(pc["id"], office, serial=serial),
    )
    assert response.status_code == 200, response.text
    return response.json()


async def _employee(
    client: AsyncClient,
    office: str,
    *,
    status: str = "Active",
) -> dict:
    tag = unique_tag()
    payload: dict[str, object] = {
        "full_name": f"Asset Employee {tag}",
        "employee_code": f"EMP-AS-{tag}",
        "email": f"asset-{tag}@example.com",
        "mobile": f"+97155{_digits(8)}",
        "designation_id": await designation_id(client),
        "employment_status": status,
        "joining_date": "2026-01-01",
        "office_id": office,
    }
    if status in {"Resigned", "Terminated"}:
        payload["last_working_date"] = "2026-08-30"
    response = await client.post("/api/v1/users", json=payload)
    assert response.status_code == 200, response.text
    return response.json()


async def _asset_admin_type(client: AsyncClient, permissions: list[str]) -> str:
    tag = unique_tag()[:8].upper()
    code = f"A{tag}"
    created = await client.post(
        "/api/v1/user-types",
        json={"name": f"Asset Admin {tag}", "code": code},
    )
    assert created.status_code == 200, created.text
    type_id = created.json()["id"]
    assert (await client.post(f"/api/v1/user-types/{type_id}/activate")).status_code == 200
    assigned = await client.put(
        f"/api/v1/user-types/{type_id}/permissions",
        json={"permissions": permissions},
    )
    assert assigned.status_code == 200, assigned.text
    scoped = await client.put(
        f"/api/v1/user-types/{type_id}/scope",
        json={"visibility_scope": "company"},
    )
    assert scoped.status_code == 200, scoped.text
    return code


@pytest.mark.asyncio
async def test_categories_asset_codes_identifiers_and_future_fields(client: AsyncClient) -> None:
    owner, _ = await owner_client(client)
    dxb = await office_id(owner, "DXB")
    categories = (await owner.get("/api/v1/assets/categories")).json()["items"]
    assert {row["name"] for row in categories} >= {
        "PC / Computer",
        "Mobile Phone",
        "SIM Card",
    }
    tag = unique_tag()[:8]
    future = await owner.post(
        "/api/v1/assets/categories",
        json={
            "code": f"TAB-{tag}",
            "name": f"Tablet {tag}",
            "description": "Future individually tracked category",
            "fields": [
                {"key": "manufacturer", "label": "Manufacturer", "required": True},
                {"key": "inventory_tag", "label": "Inventory Tag", "required": False},
            ],
        },
    )
    assert future.status_code == 200, future.text
    future_id = future.json()["id"]
    created_future = await owner.post(
        "/api/v1/assets",
        json={
            "category_id": future_id,
            "office_id": dxb,
            "condition": "Good",
            "attributes": {"manufacturer": "Framework", "inventory_tag": f"T-{tag}"},
        },
    )
    assert created_future.status_code == 200, created_future.text
    assert created_future.json()["attributes"]["manufacturer"] == "Framework"
    locked_fields = await owner.patch(
        f"/api/v1/assets/categories/{future_id}",
        json={"fields": [{"key": "other", "label": "Other", "required": False}]},
    )
    assert locked_fields.status_code == 409

    serial = f"STABLE-{unique_tag()}"
    first = await _create_pc(owner, dxb, serial=serial)
    second = await _create_pc(owner, dxb)
    assert first["assetCode"].startswith("AST-")
    assert int(second["assetCode"].split("-")[1]) == int(first["assetCode"].split("-")[1]) + 1
    duplicate = await owner.post(
        "/api/v1/assets",
        json=_pc_payload((await _category(owner, "PC"))["id"], dxb, serial=serial.lower()),
    )
    assert duplicate.status_code == 409
    assert duplicate.json()["error"]["code"] == "ASSET_SERIAL_DUPLICATE"
    immutable = await owner.patch(
        f"/api/v1/assets/{first['id']}",
        json={"asset_code": "AST-999999"},
    )
    assert immutable.status_code == 422
    corrected = await owner.post(
        f"/api/v1/assets/{first['id']}/identifiers",
        json={"serial_number": f"FIX-{unique_tag()}", "reason": "Correct service tag"},
    )
    assert corrected.status_code == 200, corrected.text
    assert corrected.json()["assetCode"] == first["assetCode"]

    deactivated = await owner.post(f"/api/v1/assets/categories/{future_id}/deactivate")
    assert deactivated.status_code == 200
    blocked = await owner.post(
        "/api/v1/assets",
        json={
            "category_id": future_id,
            "office_id": dxb,
            "condition": "New",
            "attributes": {"manufacturer": "Blocked"},
        },
    )
    assert blocked.status_code == 422
    assert (await owner.post(f"/api/v1/assets/categories/{future_id}/activate")).status_code == 200


@pytest.mark.asyncio
async def test_allocation_eligibility_return_status_and_history(client: AsyncClient) -> None:
    owner, _ = await owner_client(client)
    dxb = await office_id(owner, "DXB")
    eligible = {
        status: await _employee(owner, dxb, status=status)
        for status in ("Active", "Probation", "Notice Period")
    }
    ineligible = {
        status: await _employee(owner, dxb, status=status)
        for status in ("Resigned", "Terminated", "Inactive")
    }
    for status, employee in eligible.items():
        asset = await _create_pc(owner, dxb)
        allocated = await owner.post(
            f"/api/v1/assets/{asset['id']}/allocate",
            json={
                "employee_id": employee["id"],
                "issue_date": "2026-08-01",
                "condition_at_issue": "Good",
                "remarks": f"Allowed for {status}",
            },
        )
        assert allocated.status_code == 200, allocated.text
        assert allocated.json()["status"] == "Allocated"
        duplicate = await owner.post(
            f"/api/v1/assets/{asset['id']}/allocate",
            json={
                "employee_id": eligible["Active"]["id"],
                "issue_date": "2026-08-02",
                "condition_at_issue": "Good",
            },
        )
        assert duplicate.status_code == 409
        returned = await owner.post(
            f"/api/v1/assets/{asset['id']}/return",
            json={
                "return_date": "2026-08-20",
                "return_condition": "Good",
                "remarks": "Explicit return",
            },
        )
        assert returned.status_code == 200, returned.text
        assert returned.json()["status"] == "In Stock"
        assert returned.json()["currentAllocation"] is None
        history = await owner.get(f"/api/v1/assets/{asset['id']}/history")
        assert history.status_code == 200, history.text
        assert history.json()["allocations"][0]["returnDate"] == "2026-08-20"
    for status, employee in ineligible.items():
        asset = await _create_pc(owner, dxb)
        denied = await owner.post(
            f"/api/v1/assets/{asset['id']}/allocate",
            json={
                "employee_id": employee["id"],
                "issue_date": "2026-08-01",
                "condition_at_issue": "Good",
            },
        )
        assert denied.status_code == 422, (status, denied.text)
        assert denied.json()["error"]["code"] == "ASSET_EMPLOYEE_INELIGIBLE"

    status_asset = await _create_pc(owner, dxb)
    for status in ("Under Repair", "Damaged", "Lost", "In Stock", "Retired"):
        changed = await owner.post(
            f"/api/v1/assets/{status_asset['id']}/status",
            json={"status": status, "reason": f"Move to {status}"},
        )
        assert changed.status_code == 200, changed.text
    assert changed.json()["status"] == "Retired"


@pytest.mark.asyncio
async def test_employee_and_office_transfers_preserve_custody_chain(client: AsyncClient) -> None:
    owner, _ = await owner_client(client)
    dxb = await office_id(owner, "DXB")
    auh = await office_id(owner, "AUH")
    first_employee = await _employee(owner, dxb)
    second_employee = await _employee(owner, dxb)
    asset = await _create_pc(owner, dxb)
    assert (
        await owner.post(
            f"/api/v1/assets/{asset['id']}/allocate",
            json={
                "employee_id": first_employee["id"],
                "issue_date": "2026-07-01",
                "condition_at_issue": "Good",
            },
        )
    ).status_code == 200
    transfer = await owner.post(
        f"/api/v1/assets/{asset['id']}/transfer/employee",
        json={
            "employee_id": second_employee["id"],
            "transfer_date": "2026-07-15",
            "condition": "Good",
            "remarks": "Team handover",
        },
    )
    assert transfer.status_code == 200, transfer.text
    assert transfer.json()["currentAllocation"]["employeeId"] == second_employee["id"]
    history = (await owner.get(f"/api/v1/assets/{asset['id']}/history")).json()
    assert len(history["allocations"]) == 2
    assert {row["endType"] for row in history["allocations"]} == {
        None,
        "employee_transfer",
    }
    office_transfer_date = next(
        row["startedOn"] for row in history["officeCustody"] if row["active"]
    )
    assert (
        await owner.post(
            f"/api/v1/assets/{asset['id']}/return",
            json={"return_date": "2026-07-20", "return_condition": "Fair"},
        )
    ).status_code == 200
    office_transfer = await owner.post(
        f"/api/v1/assets/{asset['id']}/transfer/office",
        json={"office_id": auh, "transfer_date": office_transfer_date, "remarks": "AUH stock"},
    )
    assert office_transfer.status_code == 200, office_transfer.text
    assert office_transfer.json()["office"]["id"] == auh
    office_history = (await owner.get(f"/api/v1/assets/{asset['id']}/history")).json()[
        "officeCustody"
    ]
    assert len(office_history) == 2
    assert sum(1 for row in office_history if row["active"]) == 1


@pytest.mark.asyncio
async def test_offboarding_profile_reports_and_exports(client: AsyncClient) -> None:
    owner, _ = await owner_client(client)
    dxb = await office_id(owner, "DXB")
    employee = await _employee(owner, dxb)
    asset = await _create_pc(owner, dxb)
    assert (
        await owner.post(
            f"/api/v1/assets/{asset['id']}/allocate",
            json={
                "employee_id": employee["id"],
                "issue_date": "2026-08-01",
                "condition_at_issue": "Good",
            },
        )
    ).status_code == 200
    resigned = await owner.patch(
        f"/api/v1/users/{employee['id']}",
        json={"employment_status": "Resigned", "last_working_date": "2026-08-29"},
    )
    assert resigned.status_code == 200, resigned.text
    outstanding = await owner.get("/api/v1/assets/reports/outstanding_assets")
    assert outstanding.status_code == 200, outstanding.text
    assert asset["assetCode"] in {row["Asset Code"] for row in outstanding.json()["items"]}
    detail = await owner.get(f"/api/v1/assets/{asset['id']}")
    assert detail.json()["status"] == "Allocated"
    assert detail.json()["outstanding"] is True
    profile = await owner.get(f"/api/v1/assets/employees/{employee['id']}")
    assert profile.status_code == 200, profile.text
    assert profile.json()["current"][0]["asset"]["assetCode"] == asset["assetCode"]

    report_keys = {
        "asset_register",
        "available_stock",
        "allocated_assets",
        "employee_assets",
        "office_inventory",
        "damaged_assets",
        "lost_assets",
        "under_repair_assets",
        "returned_assets",
        "asset_history",
        "outstanding_assets",
    }
    for report in report_keys:
        response = await owner.get(f"/api/v1/assets/reports/{report}")
        assert response.status_code == 200, (report, response.text)
    for fmt, content_type in (
        ("xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ("pdf", "application/pdf"),
        ("print", "text/html"),
    ):
        exported = await owner.post(
            "/api/v1/assets/reports/export",
            json={"format": fmt, "report": "asset_register"},
        )
        assert exported.status_code == 200, exported.text
        assert exported.headers["content-type"].startswith(content_type)
        if fmt == "xlsx":
            workbook = load_workbook(BytesIO(exported.content))
            assert "Asset Report" in workbook.sheetnames
    csv = await owner.post(
        "/api/v1/assets/reports/export",
        json={"format": "csv", "report": "asset_register"},
    )
    assert csv.status_code == 422
    assert (
        await owner.post(
            f"/api/v1/assets/{asset['id']}/return",
            json={"return_date": date.today().isoformat(), "return_condition": "Good"},
        )
    ).status_code == 200
    cleared = await owner.get("/api/v1/assets/reports/outstanding_assets")
    assert asset["assetCode"] not in {row["Asset Code"] for row in cleared.json()["items"]}


@pytest.mark.asyncio
async def test_concurrent_identifiers_codes_and_double_allocation_are_controlled(
    client: AsyncClient,
) -> None:
    owner, _ = await owner_client(client)
    dxb = await office_id(owner, "DXB")
    permissions = [
        "Assets.View",
        "Assets.ManageStock",
        "Assets.Allocate",
    ]
    admin_type = await _asset_admin_type(owner, permissions)
    admin_a = await create_activated_user(
        owner,
        user_type_code=admin_type,
        office_id=dxb,
    )
    admin_b = await create_activated_user(
        owner,
        user_type_code=admin_type,
        office_id=dxb,
    )
    first_employee = await _employee(owner, dxb)
    second_employee = await _employee(owner, dxb)
    pc = await _category(owner, "PC")
    same_serial = f"CONCURRENT-{unique_tag()}"
    async with await spawned_client() as left, await spawned_client() as right:
        await authenticate(left, admin_a["email"], "UserPass1!")
        await authenticate(right, admin_b["email"], "UserPass1!")
        create_results = await asyncio.gather(
            left.post(
                "/api/v1/assets",
                json=_pc_payload(pc["id"], dxb, serial=same_serial),
            ),
            right.post(
                "/api/v1/assets",
                json=_pc_payload(pc["id"], dxb, serial=same_serial),
            ),
        )
        assert sorted(response.status_code for response in create_results) == [200, 409]
        assert (
            next(response for response in create_results if response.status_code == 409).json()[
                "error"
            ]["code"]
            == "ASSET_SERIAL_DUPLICATE"
        )
        winner = next(response.json() for response in create_results if response.status_code == 200)

        second_asset_results = await asyncio.gather(
            left.post("/api/v1/assets", json=_pc_payload(pc["id"], dxb)),
            right.post("/api/v1/assets", json=_pc_payload(pc["id"], dxb)),
        )
        assert all(response.status_code == 200 for response in second_asset_results)
        codes = {response.json()["assetCode"] for response in second_asset_results}
        assert len(codes) == 2

        allocation_results = await asyncio.gather(
            left.post(
                f"/api/v1/assets/{winner['id']}/allocate",
                json={
                    "employee_id": first_employee["id"],
                    "issue_date": "2026-08-01",
                    "condition_at_issue": "Good",
                },
            ),
            right.post(
                f"/api/v1/assets/{winner['id']}/allocate",
                json={
                    "employee_id": second_employee["id"],
                    "issue_date": "2026-08-01",
                    "condition_at_issue": "Good",
                },
            ),
        )
        assert sorted(response.status_code for response in allocation_results) == [200, 409]
        detail = await left.get(f"/api/v1/assets/{winner['id']}")
        assert detail.status_code == 200
        assert detail.json()["currentAllocation"]["employeeId"] in {
            first_employee["id"],
            second_employee["id"],
        }

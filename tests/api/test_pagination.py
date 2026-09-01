from __future__ import annotations

from io import BytesIO

import pytest
from helpers import (
    authenticate,
    create_activated_user,
    designation_id,
    office_id,
    owner_client,
    spawned_client,
    unique_tag,
)
from httpx import AsyncClient
from openpyxl import load_workbook


async def _create_pending_users(client: AsyncClient, count: int, marker: str) -> list[str]:
    designation = await designation_id(client)
    codes: list[str] = []
    for index in range(count):
        tag = unique_tag()
        response = await client.post(
            "/api/v1/users",
            json={
                "full_name": f"Pagination {marker} {index:03d}",
                "employee_code": f"PAG-{marker}-{index:03d}-{tag[:4]}",
                "email": f"pagination-{marker}-{index:03d}-{tag}@example.com",
                "mobile": "+971500000099",
                "designation_id": designation,
                "employment_status": "Active",
                "joining_date": "2026-02-01",
            },
        )
        assert response.status_code == 200, response.text
        codes.append(response.json()["userCode"])
    return codes


@pytest.mark.asyncio
async def test_user_pagination_boundaries_filters_and_validation(client: AsyncClient) -> None:
    owner, _ = await owner_client(client)
    marker = unique_tag()[:8]
    created_codes = await _create_pending_users(owner, 53, marker)
    query = f"q=Pagination%20{marker}"

    first = await owner.get(f"/api/v1/users?{query}")
    assert first.status_code == 200, first.text
    first_body = first.json()
    assert len(first_body["items"]) == 10
    assert first_body["pagination"] == {
        "page": 1,
        "pageSize": 10,
        "total": 53,
        "totalPages": 6,
    }
    assert [row["userCode"] for row in first_body["items"]] == sorted(created_codes)[:10]

    second = (await owner.get(f"/api/v1/users?{query}&page=2")).json()
    assert second["pagination"]["page"] == 2
    assert len(second["items"]) == 10
    assert {row["id"] for row in first_body["items"]}.isdisjoint(
        {row["id"] for row in second["items"]}
    )

    twenty_five = (await owner.get(f"/api/v1/users?{query}&page_size=25")).json()
    assert len(twenty_five["items"]) == 25
    assert twenty_five["pagination"]["totalPages"] == 3

    fifty = (await owner.get(f"/api/v1/users?{query}&page_size=50")).json()
    assert len(fifty["items"]) == 50
    final_page = (
        await owner.get(f"/api/v1/users?{query}&page=2&page_size=50")
    ).json()
    assert len(final_page["items"]) == 3
    assert final_page["pagination"]["total"] == 53

    empty = (await owner.get("/api/v1/users?q=no-such-pagination-user")).json()
    assert empty["items"] == []
    assert empty["pagination"]["total"] == 0
    assert empty["pagination"]["totalPages"] == 0

    assert (await owner.get("/api/v1/users?page=0")).status_code == 422
    assert (await owner.get("/api/v1/users?page_size=51")).status_code == 422
    assert (await owner.get("/api/v1/users?page_size=1000000")).status_code == 422


@pytest.mark.asyncio
async def test_paginated_total_does_not_leak_outside_user_scope(client: AsyncClient) -> None:
    owner, _ = await owner_client(client)
    tag = unique_tag()[:7].upper()
    user_type = await owner.post(
        "/api/v1/user-types",
        json={"name": f"Own Pagination {tag}", "code": f"OP{tag}"},
    )
    assert user_type.status_code == 200, user_type.text
    type_id = user_type.json()["id"]
    assert (await owner.post(f"/api/v1/user-types/{type_id}/activate")).status_code == 200
    permissions = await owner.put(
        f"/api/v1/user-types/{type_id}/permissions",
        json={"permissions": ["Users.View"]},
    )
    assert permissions.status_code == 200, permissions.text
    scoped = await owner.put(
        f"/api/v1/user-types/{type_id}/scope",
        json={"visibility_scope": "own"},
    )
    assert scoped.status_code == 200, scoped.text
    restricted = await create_activated_user(
        owner,
        user_type_code=user_type.json()["code"],
        password="UserPass1!",
    )

    async with await spawned_client() as restricted_client:
        await authenticate(restricted_client, restricted["email"], "UserPass1!")
        page = await restricted_client.get("/api/v1/users?page=1&page_size=50")
        assert page.status_code == 200, page.text
        assert page.json()["pagination"]["total"] == 1
        assert [row["id"] for row in page.json()["items"]] == [restricted["id"]]
        empty_page = await restricted_client.get("/api/v1/users?page=2&page_size=50")
        assert empty_page.status_code == 200, empty_page.text
        assert empty_page.json()["items"] == []
        assert empty_page.json()["pagination"]["total"] == 1


@pytest.mark.asyncio
async def test_asset_report_page_is_bounded_but_export_retains_full_scope(
    client: AsyncClient,
) -> None:
    owner, _ = await owner_client(client)
    marker = unique_tag()[:8].upper()
    category_response = await owner.post(
        "/api/v1/assets/categories",
        json={
            "code": f"PG-{marker}",
            "name": f"Pagination Assets {marker}",
            "fields": [],
        },
    )
    assert category_response.status_code == 200, category_response.text
    category_id = category_response.json()["id"]
    dxb = await office_id(owner, "DXB")
    asset_codes: set[str] = set()
    for index in range(12):
        created = await owner.post(
            "/api/v1/assets",
            json={
                "category_id": category_id,
                "office_id": dxb,
                "condition": "New",
                "description": f"Pagination export {marker} {index:02d}",
            },
        )
        assert created.status_code == 200, created.text
        asset_codes.add(created.json()["assetCode"])

    report = await owner.get(
        f"/api/v1/assets/reports/asset_register?categoryId={category_id}"
    )
    assert report.status_code == 200, report.text
    assert len(report.json()["items"]) == 10
    assert report.json()["pagination"]["total"] == 12

    exported = await owner.post(
        "/api/v1/assets/reports/export",
        json={
            "format": "xlsx",
            "report": "asset_register",
            "category_id": category_id,
        },
    )
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(BytesIO(exported.content), read_only=True)
    values = {
        str(value)
        for row in workbook["Asset Report"].iter_rows(values_only=True)
        for value in row
        if value is not None
    }
    assert asset_codes <= values
